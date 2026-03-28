"""Excel 边缘场景变体生成器

基于 PRD 真实 Excel 动态生成变体，用于测试导入容错能力。
"""
from __future__ import annotations

import copy
import random
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _load_source(source_path: str | Path) -> Workbook:
    return load_workbook(str(source_path))


def _save_tmp(wb: Workbook, suffix: str = "") -> Path:
    fd, path = tempfile.mkstemp(suffix=f"_{suffix}.xlsx", prefix="e2e_edge_")
    wb.save(path)
    wb.close()
    return Path(path)


# ---- Sheet 结构变化 ----

def with_extra_sheet(source: str | Path) -> Path:
    """添加额外 sheet（2026年1月批零）"""
    wb = _load_source(source)
    # 复制第一个 sheet 并重命名
    src_ws = wb.worksheets[0]
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = "2026年1月批零"
    return _save_tmp(wb, "extra_sheet")


def with_missing_sheet(source: str | Path, sheet_name: str = "吃穿用") -> Path:
    """删除指定 sheet"""
    wb = _load_source(source)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    return _save_tmp(wb, "missing_sheet")


def with_shuffled_sheets(source: str | Path) -> Path:
    """随机打乱 sheet 顺序"""
    wb = _load_source(source)
    names = wb.sheetnames[:]
    random.shuffle(names)
    for i, name in enumerate(names):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    return _save_tmp(wb, "shuffled_sheets")


def with_sheet_name_spaces(source: str | Path) -> Path:
    """Sheet 名添加前后空格"""
    wb = _load_source(source)
    for ws in wb.worksheets[:3]:
        ws.title = f" {ws.title}  "
    return _save_tmp(wb, "sheet_spaces")


def with_duplicate_type_sheet(source: str | Path) -> Path:
    """复制批发 sheet 创建重复类型"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            dup = wb.copy_worksheet(ws)
            dup.title = "批发_副本"
            break
    return _save_tmp(wb, "dup_type_sheet")


# ---- 列结构变化 ----

def with_extra_empty_columns(source: str | Path) -> Path:
    """在主工作表中间插入空列"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            ws.insert_cols(3, 2)  # 在第 3 列后插入 2 列
            break
    return _save_tmp(wb, "extra_cols")


def with_swapped_columns(source: str | Path) -> Path:
    """交换两列位置"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            # 交换第 4、5 列的值（跳过合并单元格）
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 300)):
                cells = list(row)
                if len(cells) >= 5:
                    try:
                        v3, v4 = cells[3].value, cells[4].value
                        cells[3].value = v4
                        cells[4].value = v3
                    except AttributeError:
                        pass  # 跳过 MergedCell
            break
    return _save_tmp(wb, "swapped_cols")


def with_column_name_spaces(source: str | Path) -> Path:
    """列名添加多余空格"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str):
                    cell.value = f"  {cell.value}  "
            break
    return _save_tmp(wb, "col_spaces")


def with_column_name_variant(source: str | Path) -> Path:
    """列名分号替换为冒号"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value and isinstance(cell.value, str) and ";" in cell.value:
                cell.value = cell.value.replace(";", "：")
    return _save_tmp(wb, "col_variant")


def with_missing_non_core_column(source: str | Path) -> Path:
    """删除非核心列（粮油食品类）"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and "粮油" in str(cell.value):
                    ws.delete_cols(idx)
                    break
            break
    return _save_tmp(wb, "missing_col")


# ---- 数据变化 ----

def with_empty_rows(source: str | Path) -> Path:
    """在中间和末尾插入空行"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            ws.insert_rows(5)  # 中间插入空行
            if ws.max_row:
                ws.insert_rows(ws.max_row + 1)  # 末尾插入空行
            break
    return _save_tmp(wb, "empty_rows")


def with_text_numbers(source: str | Path) -> Path:
    """将部分数值存为文本格式"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 1, 10)):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.value = str(cell.value)
                        break  # 每行只改一个
            break
    return _save_tmp(wb, "text_numbers")


def with_comma_numbers(source: str | Path) -> Path:
    """数值带千分位逗号"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 1, 10)):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 1000:
                        cell.value = f"{cell.value:,.2f}"
                        break
            break
    return _save_tmp(wb, "comma_numbers")


def with_null_cells(source: str | Path) -> Path:
    """清空部分数值单元格"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row or 1, 8)):
                cells = list(row)
                if len(cells) > 5:
                    cells[5].value = None
            break
    return _save_tmp(wb, "null_cells")


def with_extreme_values(source: str | Path) -> Path:
    """设置极端大数值"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=2, max_row=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.value = 9999999999.99
                        break
            break
    return _save_tmp(wb, "extreme_values")


def with_zero_rows(source: str | Path) -> Path:
    """企业所有金额为 0"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=2, max_row=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.value = 0
            break
    return _save_tmp(wb, "zero_rows")


def with_bad_credit_code(source: str | Path) -> Path:
    """信用代码格式异常（非 18 位）"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            for row in ws.iter_rows(min_row=2, max_row=4):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and len(cell.value) == 18:
                        cell.value = "INVALID_CODE"
                        break
            break
    return _save_tmp(wb, "bad_credit")


# ---- 日期/月份变化 ----

def with_different_month(source: str | Path) -> Path:
    """列名中月份替换为 1 月"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value and isinstance(cell.value, str) and "12月" in cell.value:
                cell.value = cell.value.replace("12月", "1月")
    return _save_tmp(wb, "diff_month")


def with_different_year(source: str | Path) -> Path:
    """列名中年份替换"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value and isinstance(cell.value, str) and "2025年" in cell.value:
                cell.value = cell.value.replace("2025年", "2026年")
    return _save_tmp(wb, "diff_year")


# ---- 整体文件异常 ----

def empty_excel() -> Path:
    """空 Excel（只有一个空 sheet）"""
    wb = Workbook()
    return _save_tmp(wb, "empty")


def single_sheet_only(source: str | Path) -> Path:
    """只保留第一个 sheet"""
    wb = _load_source(source)
    while len(wb.sheetnames) > 1:
        wb.remove(wb.worksheets[-1])
    return _save_tmp(wb, "single_sheet")


def headers_only(source: str | Path) -> Path:
    """只有表头没有数据"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if ws.max_row and ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    return _save_tmp(wb, "headers_only")


def large_file(source: str | Path, target_rows: int = 10000) -> Path:
    """复制企业行生成大文件"""
    wb = _load_source(source)
    for ws in wb.worksheets:
        if "批发" in ws.title and "批零" not in ws.title:
            if ws.max_row and ws.max_row > 2:
                # 复制第 2 行数据
                src_row = [cell.value for cell in ws[2]]
                current = ws.max_row
                for i in range(current + 1, min(current + target_rows, current + 500)):
                    for col, val in enumerate(src_row, 1):
                        ws.cell(row=i, column=col, value=val)
            break
    return _save_tmp(wb, "large_file")
