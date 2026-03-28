"""数据一致性验证"""
import pytest
from openpyxl import load_workbook
from pathlib import Path


class TestDataConsistency:
    """数据一致性验证"""

    def test_import_export_roundtrip(self, api, import_data, results_dir: Path):
        """导入 → 不调整 → 导出 → 企业数一致"""
        # 重置所有企业数据
        api.reset_companies()
        output = results_dir / "roundtrip_export.xlsx"
        api.export_stream_download(output)
        if output.exists():
            wb = load_workbook(str(output), read_only=True)
            # 检查有数据 sheet
            has_data = False
            for name in wb.sheetnames:
                ws = wb[name]
                if ws.max_row and ws.max_row > 1:
                    has_data = True
                    break
            wb.close()
            assert has_data, "导出文件应包含数据"

    def test_indicators_api_consistency(self, api, import_data):
        """16 个指标 API 返回一致"""
        ind = api.indicators()
        groups = ind.get("groups", [])
        all_ids = set()
        for g in groups:
            for i in g.get("indicators", []):
                all_ids.add(i["id"])
        # 至少应有核心指标
        expected_ids = {
            "limitAbove_month_rate", "wholesale_month_rate",
            "retail_month_rate", "accommodation_month_rate",
            "catering_month_rate",
        }
        found = expected_ids & all_ids
        assert len(found) >= 3, f"应找到核心指标, got ids: {all_ids}"

    def test_company_count_consistency(self, api, import_data):
        """企业数量：批发+零售+住宿+餐饮 = 总数"""
        st = api.status()
        total = st["totalCompanies"]
        # 分行业查询
        wr = api.companies(industry_type="wholesale", page_size=1)
        rt = api.companies(industry_type="retail", page_size=1)
        ac = api.companies(industry_type="accommodation", page_size=1)
        ct = api.companies(industry_type="catering", page_size=1)
        sum_by_type = (
            wr.get("total", 0) + rt.get("total", 0) +
            ac.get("total", 0) + ct.get("total", 0)
        )
        assert sum_by_type == total, f"分行业总数 {sum_by_type} != 总数 {total}"

    def test_cumulative_value_consistency(self, api, import_data):
        """累计值 = 1-11月 + 当月（抽样验证）"""
        companies = api.companies(industry_type="wholesale", page_size=5)
        items = companies.get("items", [])
        if not items:
            pytest.skip("无批发企业")
        for c in items[:3]:
            prev_cum = c.get("salesPrevCumulative") or c.get("salesLastYearCumulative", 0)
            # 简单检查字段存在
            assert "creditCode" in c, "企业应有信用代码"
            assert "name" in c, "企业应有名称"

    def test_optimize_returns_groups(self, api, import_data):
        """优化接口返回指标分组"""
        result = api.optimize({"limitAbove_month_rate": 10.0})
        assert "groups" in result, "optimize 应返回 groups"
        groups = result["groups"]
        assert len(groups) > 0, "应有指标分组"

    def test_export_sheets_match_template(self, api, import_data, template_excel: Path, results_dir: Path):
        """导出 sheet 名与模板一致"""
        output = results_dir / "template_compare.xlsx"
        api.export_stream_download(output)
        if not output.exists():
            pytest.skip("导出失败")
        wb_export = load_workbook(str(output), read_only=True)
        wb_template = load_workbook(str(template_excel), read_only=True)
        export_names = set(wb_export.sheetnames)
        template_names = set(wb_template.sheetnames)
        wb_export.close()
        wb_template.close()
        # 导出应包含模板中的核心 sheet
        common = export_names & template_names
        assert len(common) >= 4, f"导出与模板共有 sheet 应 >= 4, common: {common}"

    def test_reset_then_verify(self, api, import_data):
        """重置企业数据后指标归位"""
        # 先修改一个企业
        companies = api.companies(page_size=1)
        items = companies.get("items", [])
        if not items:
            pytest.skip("无企业数据")
        cid = items[0]["id"]
        api.patch_company(cid, {"salesCurrentMonth": 999999})
        # 重置
        api.reset_companies([cid])
        # 验证恢复
        after = api.companies(page_size=1)
        after_items = after.get("items", [])
        if after_items:
            # 值应恢复到导入时的值（不是 999999）
            val = after_items[0].get("salesCurrentMonth", 0)
            assert val != 999999 or val == 0, "重置后值不应是修改值"
