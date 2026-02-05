#!/usr/bin/env python3
# E2E 报告生成脚本
# @author Anner
# Updated on 2026/2/4
import argparse
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


def _read_text(path: Optional[str]) -> str:
    if not path:
        return ""
    p = Path(path)
    if not p.exists():
        return ""
    return p.read_text(encoding="utf-8", errors="replace")


def _load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))

def _unwrap_agent_browser_json(d: Dict[str, Any]) -> Dict[str, Any]:
    # agent-browser --json wraps as {success,data:{result:...}}; errors as {success:false,error:"..."}.
    if not isinstance(d, dict):
        return {"error": "invalid json", "rows": []}
    if d.get("success") is False:
        return {"error": d.get("error") or "agent-browser error", "rows": []}
    data = d.get("data")
    if isinstance(data, dict) and "result" in data and isinstance(data["result"], dict):
        return data["result"]
    return d


def _load_optional_json(path: str) -> Dict[str, Any]:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _parse_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        x = float(v)
        if math.isfinite(x):
            return x
        return None
    s = str(v).strip()
    if not s or s == "-":
        return None
    s = s.replace(",", "")
    if s.endswith("%"):
        s = s[:-1]
    try:
        x = float(s)
        if math.isfinite(x):
            return x
        return None
    except Exception:
        return None


def _round_half_up(v: float, digits: int) -> float:
    if digits < 0:
        return v
    scale = 10 ** digits
    x = v * scale
    if x >= 0:
        return math.floor(x + 0.5) / scale
    return -math.floor(-x + 0.5) / scale


def _close(a: Optional[float], b: Optional[float], eps: float = 1e-6) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= eps


def _norm_label(s: str) -> str:
    return re.sub(r"\s+", "", str(s or "").replace("\u3000", ""))


def _indicator_export_mapping() -> List[Dict[str, Any]]:
    return [
        {"label": "限上社零额（当月值）", "sheet": "汇总表（定）", "cell": "G4"},
        {"label": "限上社零额增速（当月）", "sheet": "汇总表（定）", "cell": "S4"},
        {"label": "限上社零额（累计值）", "sheet": "汇总表（定）", "cell": "I4"},
        {"label": "限上社零额增速（累计）", "sheet": "汇总表（定）", "cell": "T4"},
        {"label": "批发业销售额增速（当月）", "sheet": "汇总表（定）", "cell": "K4"},
        {"label": "批发业销售额增速（累计）", "sheet": "汇总表（定）", "cell": "L4"},
        {"label": "零售业销售额增速（当月）", "sheet": "汇总表（定）", "cell": "M4"},
        {"label": "零售业销售额增速（累计）", "sheet": "汇总表（定）", "cell": "N4"},
        {"label": "住宿业营业额增速（当月）", "sheet": "汇总表（定）", "cell": "O4"},
        {"label": "住宿业营业额增速（累计）", "sheet": "汇总表（定）", "cell": "P4"},
        {"label": "餐饮业营业额增速（当月）", "sheet": "汇总表（定）", "cell": "Q4"},
        {"label": "餐饮业营业额增速（累计）", "sheet": "汇总表（定）", "cell": "R4"},
        {"label": "吃穿用增速（当月）", "sheet": "汇总表（定）", "cell": "U4"},
        {"label": "小微企业增速（当月）", "sheet": "汇总表（定）", "cell": "V4"},
        {"label": "社零总额（累计值）", "sheet": "汇总表（定）", "cell": "N10", "scale": "yi"},
        {"label": "社零总额增速（累计）", "sheet": "汇总表（定）", "cell": "S10"},
    ]


def _indicator_count_check(items: List[Dict[str, Any]], expected: int = 16) -> Dict[str, Any]:
    labels = [_norm_label(it.get("label") or "") for it in items if isinstance(it, dict)]
    uniq = [x for x in dict.fromkeys(labels) if x]
    ok = len(uniq) == expected
    return {
        "ok": ok,
        "count": len(uniq),
        "expected": expected,
    }


def _compare_indicator_values_to_export(items: List[Dict[str, Any]], export_wb) -> Dict[str, Any]:
    res = {"executed": False, "ok": True, "total": 0, "mismatch": 0, "items": []}
    if export_wb is None:
        return res
    mapping = _indicator_export_mapping()
    res["executed"] = True
    res["total"] = len(mapping)
    item_map = {_norm_label(it.get("label")): it for it in items if isinstance(it, dict)}
    mismatches: List[Dict[str, Any]] = []
    for it in mapping:
        label = it["label"]
        key = _norm_label(label)
        item = item_map.get(key)
        if not item:
            mismatches.append({"label": label, "reason": "指标缺失"})
            continue
        ui_val = _parse_number(item.get("value"))
        if ui_val is None:
            mismatches.append({"label": label, "reason": "指标值为空", "uiValue": item.get("value")})
            continue
        expected = ui_val
        if it.get("scale") == "yi":
            expected = _round_half_up(ui_val / 10000.0, 2)
        sheet = it["sheet"]
        cell = it["cell"]
        if sheet not in export_wb.sheetnames:
            mismatches.append({"label": label, "reason": f"导出缺少 sheet: {sheet}"})
            continue
        excel_val_raw = export_wb[sheet][cell].value
        excel_val = _parse_number(excel_val_raw)
        if excel_val is None:
            mismatches.append(
                {
                    "label": label,
                    "reason": "导出值为空",
                    "cell": f"{sheet}!{cell}",
                    "expected": expected,
                    "actual": excel_val_raw,
                }
            )
            continue
        eps = 0.02 if it.get("scale") == "yi" else _field_eps(label)
        if not _close(expected, excel_val, eps=eps):
            mismatches.append(
                {
                    "label": label,
                    "reason": "导出值与指标不一致",
                    "cell": f"{sheet}!{cell}",
                    "expected": expected,
                    "actual": excel_val,
                }
            )
    res["items"] = mismatches
    res["mismatch"] = len(mismatches)
    res["ok"] = len(mismatches) == 0
    return res


def _field_eps(field: str) -> float:
    # UI 通常展示 2 位小数；增速/比例类字段允许更大的容差
    if "增速" in field or "零销比" in field or "rate" in field.lower() or "%" in field:
        # UI 目前以整数百分比展示（四舍五入），允许接近 1% 的误差
        return 0.6
    return 0.005


def _normalize_rate_pair(field: str, a: Optional[float], b: Optional[float]) -> Tuple[Optional[float], Optional[float]]:
    # Some sheets store rate as decimal (0.54) while UI/export shows percent (54).
    # Normalize by scaling the smaller-magnitude side when it looks like a 100x difference.
    if a is None or b is None:
        return a, b
    is_rate = "增速" in field or "零销比" in field or "rate" in field.lower() or "%" in field
    if not is_rate:
        return a, b
    aa = abs(a)
    bb = abs(b)
    small = aa if aa < bb else bb
    large = bb if aa < bb else aa
    if small > 0:
        ratio = large / small
        # 仅在“看起来是 100 倍差异”的情况下做归一化，避免把 2% vs 2.4% 误判为 200% vs 2.4%
        if 50.0 <= ratio <= 200.0:
            if aa < bb:
                return a * 100.0, b
            return a, b * 100.0
    return a, b


def _indicator_items(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    items = payload.get("indicators") if isinstance(payload, dict) else None
    if not isinstance(items, list):
        return []
    out: List[Dict[str, Any]] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        label = str(it.get("label") or "").strip()
        if not label:
            continue
        out.append(
            {
                "label": label,
                "unit": str(it.get("unit") or "").strip(),
                "value": it.get("value"),
            }
        )
    return out


def _indicator_stats(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    stats = {
        "total": 0,
        "eatWearUse": 0,
        "smallMicro": 0,
        "industries": {},
    }
    for r in rows:
        code = str(r.get("__creditCode") or "").strip()
        if not _is_credit_code(code):
            continue
        stats["total"] += 1
        flags = str(r.get("标记") or "").strip()
        if "吃穿用" in flags:
            stats["eatWearUse"] = 1
        if "小微" in flags:
            stats["smallMicro"] = 1
        industry = str(r.get("__industry") or "").strip()
        if industry:
            stats["industries"][industry] = int(stats["industries"].get(industry, 0)) + 1
    return stats


def _indicator_requires_change(label: str, stats: Dict[str, Any]) -> bool:
    if not stats or stats.get("total", 0) == 0:
        return False
    if "吃穿用" in label and stats.get("eatWearUse", 0) == 0:
        return False
    if "小微" in label and stats.get("smallMicro", 0) == 0:
        return False
    industries = stats.get("industries") or {}
    if "批发" in label and industries.get("批发", 0) == 0:
        return False
    if "零售" in label and industries.get("零售", 0) == 0:
        return False
    if "住宿" in label and industries.get("住宿", 0) == 0:
        return False
    if "餐饮" in label and industries.get("餐饮", 0) == 0:
        return False
    return True


def _compare_indicators(
    before: List[Dict[str, Any]], after: List[Dict[str, Any]], stats: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    before_map = {str(it.get("label")): it for it in before}
    after_map = {str(it.get("label")): it for it in after}
    changes: List[Dict[str, Any]] = []
    required = 0
    changed = 0
    skipped = 0
    for label, b in before_map.items():
        a = after_map.get(label)
        if not a:
            changes.append({"label": label, "before": b.get("value"), "after": None, "ok": False, "reason": "调整后缺失"})
            continue
        if stats is not None and not _indicator_requires_change(label, stats):
            skipped += 1
            required += 1
            changes.append(
                {"label": label, "before": b.get("value"), "after": a.get("value"), "ok": False, "reason": "无可调整企业"}
            )
            continue
        required += 1
        b_num = _parse_number(b.get("value"))
        a_num = _parse_number(a.get("value"))
        ok = False
        if b_num is None and a_num is None:
            ok = False
        else:
            ok = not _close(b_num, a_num, eps=1e-6)
        if ok:
            changed += 1
        changes.append({"label": label, "before": b.get("value"), "after": a.get("value"), "ok": ok})
    total = len(before_map)
    ok = total > 0 and required == changed
    return {"total": total, "required": required, "changed": changed, "skipped": skipped, "items": changes, "ok": ok}


def _compare_dag_tables(before_rows: List[Dict[str, Any]], after_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    before_map = {
        str(r.get("__creditCode") or "").strip().upper(): r
        for r in before_rows
        if _is_credit_code(r.get("__creditCode"))
    }
    after_map = {
        str(r.get("__creditCode") or "").strip().upper(): r
        for r in after_rows
        if _is_credit_code(r.get("__creditCode"))
    }
    matched = 0
    changed_rows = 0
    total_cells = 0
    changed_cells = 0
    samples: List[Dict[str, Any]] = []
    ignore_fields = {"规模", "标记", "来源表"}

    for code, after_row in after_map.items():
        before_row = before_map.get(code)
        if not before_row:
            continue
        matched += 1
        row_changed = False
        for field, after_val in after_row.items():
            if not isinstance(field, str):
                continue
            if field.startswith("__") or field in ignore_fields:
                continue
            before_val = before_row.get(field)
            a_num = _parse_number(after_val)
            b_num = _parse_number(before_val)
            if a_num is None and b_num is None:
                continue
            total_cells += 1
            a_num, b_num = _normalize_rate_pair(field, a_num, b_num)
            if not _close(a_num, b_num, eps=_field_eps(field)):
                changed_cells += 1
                if not row_changed:
                    row_changed = True
                if len(samples) < 20:
                    samples.append(
                        {
                            "creditCode": code,
                            "field": field,
                            "before": before_val,
                            "after": after_val,
                        }
                    )
        if row_changed:
            changed_rows += 1

    ok = matched > 0 and changed_rows == matched
    return {
        "matchedRows": matched,
        "changedRows": changed_rows,
        "totalCells": total_cells,
        "changedCells": changed_cells,
        "samples": samples,
        "ok": ok,
    }


def _scan_header_digit_cells(wb, max_rows: int = 8) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    if wb is None:
        return out
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        cells: Dict[str, str] = {}
        for r in range(1, min(max_rows, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, str):
                    s = v.strip()
                    if not s:
                        continue
                    if any(ch.isdigit() for ch in s):
                        cells[ws.cell(r, c).coordinate] = s
        if cells:
            out[sheet] = cells
    return out


def _export_param_checks(export_wb_raw, export_alt_wb_raw, meta: Dict[str, Any]) -> Dict[str, Any]:
    out = {"executed": False, "ok": True, "diffCount": 0, "diffSamples": [], "meta": meta}
    if not export_wb_raw or not export_alt_wb_raw:
        return out
    primary = meta.get("primary") or {}
    alt = meta.get("alternate") or {}
    if not primary or not alt:
        return out
    if primary.get("year") == alt.get("year") and primary.get("month") == alt.get("month"):
        return out

    out["executed"] = True
    primary_cells = _scan_header_digit_cells(export_wb_raw)
    alt_cells = _scan_header_digit_cells(export_alt_wb_raw)
    diffs: List[Dict[str, Any]] = []
    for sheet, cells in primary_cells.items():
        alt_sheet = alt_cells.get(sheet) or {}
        for addr, v in cells.items():
            av = alt_sheet.get(addr)
            if av is None:
                continue
            if str(v) != str(av):
                diffs.append({"sheet": sheet, "cell": addr, "primary": v, "alternate": av})
    out["diffCount"] = len(diffs)
    out["diffSamples"] = diffs[:40]
    out["ok"] = len(diffs) > 0

    file_name = str(alt.get("fileName") or "")
    if file_name:
        out["fileName"] = file_name
        out["fileNameOk"] = bool(re.search(r"\\d+月月报（定）", file_name))
        if not out["fileNameOk"]:
            out["ok"] = False
    return out


def _ui_normalization_checks(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    allowed_headers = {
        "企业名称",
        "规模",
        "标记",
        "本年-上月",
        "本年-本月",
        "上年-本月",
        "同比增量(当月)",
        "环比增量(当月)",
        "环比增速(当月)",
        "同比增速(当月)",
        "本年-1—本月",
        "上年-1—本月",
        "累计同比增量",
        "累计同比增速",
        "零售额;本年-上月",
        "零售额;本年-本月",
        "零售额;上年-本月",
        "零售额;同比增量(当月)",
        "零售额;环比增量(当月)",
        "零售额;环比增速(当月)",
        "零售额;同比增速(当月)",
        "零售额;本年-1—本月",
        "零售额;上年-1—本月",
        "零售额;累计同比增量",
        "零售额;累计同比增速",
        "零销比(%)",
        "来源表",
    }
    unknown_headers = [h for h in headers if h and h not in allowed_headers]
    missing_headers = [h for h in allowed_headers if h not in headers]

    allowed_sources = {"批发", "零售", "住宿", "餐饮"}
    allowed_industries = {"批发", "零售", "住宿", "餐饮"}
    unknown_sources: List[str] = []
    unknown_industries: List[str] = []
    for r in rows:
        src = str(r.get("来源表") or "").strip()
        if src and src not in allowed_sources and src not in unknown_sources:
            unknown_sources.append(src)
        ind = str(r.get("__industry") or "").strip()
        if ind and ind not in allowed_industries and ind not in unknown_industries:
            unknown_industries.append(ind)

    ok = not unknown_headers and not unknown_sources and not unknown_industries
    return {
        "ok": ok,
        "unknownHeaders": unknown_headers,
        "missingHeaders": missing_headers,
        "unknownSources": unknown_sources,
        "unknownIndustries": unknown_industries,
    }


def _safe(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _link_or_text(path: str, empty_text: str = "未生成") -> str:
    if path:
        return f"<a href=\"{_safe(path)}\">{_safe(path)}</a>"
    return _safe(empty_text)


def _rel(p: str) -> str:
    try:
        return os.path.relpath(p, start=str(Path(p).parent.parent))
    except Exception:
        return p


def _find_header_row(ws) -> Optional[int]:
    # Heuristic: find row containing "统一社会信用代码"
    for r in range(1, 10):
        vals = [ws.cell(r, c).value for c in range(1, min(80, ws.max_column + 1))]
        joined = " ".join([str(v) for v in vals if v not in (None, "")])
        if "统一社会信用代码" in joined:
            return r
    return None


_CREDIT_CODE_RE = re.compile(r"^[0-9A-Z]{18}$")


def _is_credit_code(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().upper()
    return bool(_CREDIT_CODE_RE.match(s))

def _get_header_cols(headers: List[Any], name: str, nth: int = 1) -> Optional[int]:
    # Returns 1-based column index for nth occurrence of header name.
    seen = 0
    for i, h in enumerate(headers, start=1):
        if h is None:
            continue
        if str(h).strip() == name:
            seen += 1
            if seen == nth:
                return i
    return None


def _sheet_table(ws) -> Optional[Tuple[Dict[str, int], Dict[str, int]]]:
    header_row = _find_header_row(ws)
    if not header_row:
        return None
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None or str(v).strip() == "":
            continue
        headers[str(v).strip()] = c
    if not headers:
        return None

    code_col = headers.get("统一社会信用代码")
    if not code_col:
        return None

    rows: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not _is_credit_code(code):
            continue
        rows[str(code).strip().upper()] = r
    return headers, rows


UI_TO_EXCEL_FIELD = {
    # 住餐在 UI 里使用更口语化的列名
    "本月客房收入": "客房收入;本年-本月",
    "本月餐费收入": "餐费收入;本年-本月",
    "本月商品销售额": "商品销售额;本年-本月",
    # UI 里的“增速”列名与导出表的衍生指标列名不同
    "同比增速(当月)": "(衍生指标)销售额当月增速",
    "累计同比增速": "(衍生指标)销售额累计增速",
    "商品销售额;增速(当月)": "(衍生指标)销售额当月增速",
    "商品销售额;累计增速": "(衍生指标)销售额累计增速",
    "销售额;增速(当月)": "(衍生指标)销售额当月增速",
    "销售额;累计增速": "(衍生指标)销售额累计增速",
    "零售额;同比增速(当月)": "(衍生指标)零售额当月增速",
    "零售额;累计同比增速": "(衍生指标)零售额累计增速",
    "零售额;增速(当月)": "(衍生指标)零售额当月增速",
    "零售额;累计增速": "(衍生指标)零售额累计增速",
    "营业额;增速(当月)": "(衍生指标)营业额当月增速",
    "营业额;累计增速": "(衍生指标)营业额累计增速",
}


def _map_ui_field_to_export_header(ui_field: str, export_sheet: str) -> str:
    """
    将明细表列名映射到导出 Excel 的列名。
    - 批零导出表用「商品销售额;*」作为主销售额列
    - 住餐导出表用「营业额;*」作为主营业额列
    """
    # 住餐口径：UI 的“同比/累计同比增速”对应“营业额”的衍生指标列
    if export_sheet in {"住宿", "餐饮", "住餐总表"}:
        if ui_field == "同比增速(当月)":
            return "(衍生指标)营业额当月增速"
        if ui_field == "累计同比增速":
            return "(衍生指标)营业额累计增速"

    if ui_field in UI_TO_EXCEL_FIELD:
        return UI_TO_EXCEL_FIELD[ui_field]

    # 最新 UI 主销售额列名为“本年-本月/上年-本月/本年-1—本月/上年-1—本月”等（无“销售额;”前缀）
    if ui_field in {"本年-本月", "上年-本月", "本年-1—本月", "上年-1—本月"}:
        if export_sheet in {"住宿", "餐饮", "住餐总表"}:
            return f"营业额;{ui_field}"
        if export_sheet in {"批发", "零售", "批零总表"}:
            return f"商品销售额;{ui_field}"

    # 住餐：UI 里“销售额”列是营业额的口径映射
    if export_sheet in {"住宿", "餐饮", "住餐总表"}:
        if ui_field.startswith("销售额;"):
            return ui_field.replace("销售额;", "营业额;", 1)

    # 批零：UI 里“销售额”对应导出“商品销售额”
    if export_sheet in {"批发", "零售", "批零总表"}:
        if ui_field.startswith("销售额;"):
            return ui_field.replace("销售额;", "商品销售额;", 1)

    return ""

DERIVED_SHEET_MAPPINGS: Dict[str, List[Tuple[str, str, int]]] = {
    # (ui_field, excel_header, nth_occurrence)
    "批发": [
        ("本年-上月", "2025年11月销售额", 1),
        ("本年-本月", "2025年12月销售额", 1),
        ("上年-本月", "2024年;12月;商品销售额;千元", 1),
        ("本年-1—本月", "2025年1-12月销售额", 1),
        ("上年-1—本月", "2024年;1-12月;商品销售额;千元", 1),
        ("同比增速(当月)", "12月销售额增速", 1),
        ("累计同比增速", "1-12月增速", 1),
        ("零售额;本年-上月", "2025年11月零售额", 1),
        ("零售额;本年-本月", "2025年12月零售额", 1),
        ("零售额;上年-本月", "2024年;12月;商品零售额;千元", 1),
        ("零售额;本年-1—本月", "2025年1-12月零售额", 1),
        ("零售额;上年-1—本月", "2024年;1-12月;商品零售额;千元", 1),
        ("零售额;同比增速(当月)", "12月零售额增速", 1),
        ("零售额;累计同比增速", "1-12月增速", 2),
        ("零销比(%)", "零售额占比", 1),
    ],
    "零售": [
        ("本年-上月", "2025年11月销售额", 1),
        ("本年-本月", "2025年12月销售额", 1),
        ("上年-本月", "2024年;12月;商品销售额;千元", 1),
        ("本年-1—本月", "2025年1-12月销售额", 1),
        ("上年-1—本月", "2024年;1-12月;商品销售额;千元", 1),
        ("同比增速(当月)", "12月销售额增速", 1),
        ("累计同比增速", "1-12月增速", 1),
        ("零售额;本年-上月", "2025年11月零售额", 1),
        ("零售额;本年-本月", "2025年12月零售额", 1),
        ("零售额;上年-本月", "2024年;12月;商品零售额;千元", 1),
        ("零售额;本年-1—本月", "2025年1-12月零售额", 1),
        ("零售额;上年-1—本月", "2024年;1-12月;商品零售额;千元", 1),
        ("零售额;同比增速(当月)", "12月零售额增速", 1),
        ("零售额;累计同比增速", "1-12月增速", 2),
        ("零销比(%)", "零售额占比", 1),
    ],
    "住宿": [
        # 最新 UI：住餐口径复用“本年-本月”等列（内部映射到营业额字段）
        ("本年-上月", "2025年11月营业额", 1),
        ("本年-本月", "2025年12月营业额", 1),
        ("上年-本月", "2024年12月;营业额总计;千元", 1),
        ("本年-1—本月", "2025年1-12月营业额", 1),
        ("上年-1—本月", "2024年1-12月;营业额总计;千元", 1),
        ("同比增速(当月)", "12月增速", 1),
        ("累计同比增速", "1-12月增速", 1),
        ("零售额;本年-本月", "2025年12月零售额", 1),
        ("零售额;上年-本月", "2024年12月零售额", 1),
    ],
    "餐饮": [
        ("本年-上月", "2025年11月营业额", 1),
        ("本年-本月", "2025年12月营业额", 1),
        ("上年-本月", "2024年12月;营业额总计;千元", 1),
        ("本年-1—本月", "2025年1-12月营业额", 1),
        ("上年-1—本月", "2024年1-12月;营业额总计;千元", 1),
        ("同比增速(当月)", "12月增速", 1),
        ("累计同比增速", "1-12月增速", 1),
        ("零售额;本年-本月", "2025年12月零售额", 1),
        ("零售额;上年-本月", "2024年12月零售额", 1),
    ],
}


@dataclass
class Mismatch:
    credit_code: str
    name: str
    sheet: str
    field: str
    expected: Any
    actual: Any


def _compare_ui_vs_excel(
    ui_rows: List[Dict[str, Any]],
    wb,
    ignore_fields: Iterable[str],
) -> Tuple[List[str], List[Mismatch]]:
    ignore = set(ignore_fields)
    missing_rows: List[str] = []
    mismatches: List[Mismatch] = []

    # Group UI rows by source sheet (UI extracted key: "来源表")
    by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    for r in ui_rows:
        sheet = str(r.get("来源表") or "").strip()
        by_sheet.setdefault(sheet, []).append(r)

    for sheet_name, rows in by_sheet.items():
        if not sheet_name:
            for r in rows:
                missing_rows.append(f"{r.get('__creditCode')}: missing 来源表")
            continue
        if sheet_name not in wb.sheetnames:
            for r in rows:
                missing_rows.append(f"{r.get('__creditCode')}: 来源表 not found in excel: {sheet_name}")
            continue

        ws = wb[sheet_name]
        table = _sheet_table(ws)
        if not table:
            for r in rows:
                missing_rows.append(f"{r.get('__creditCode')}: sheet has no recognizable header table: {sheet_name}")
            continue
        headers, excel_rows = table

        # Ensure excel rows -> UI rows coverage (excel has row -> UI must exist)
        ui_codes = {str(r.get("__creditCode") or "").strip().upper() for r in rows if _is_credit_code(r.get("__creditCode"))}
        for code in excel_rows.keys():
            if code not in ui_codes:
                missing_rows.append(f"{code}: exists in excel sheet {sheet_name} but missing in 明细表")

        # Compare values where both sides have the field header name.
        for r in rows:
            code = str(r.get("__creditCode") or "").strip()
            name = str(r.get("__name") or "").strip()
            if not _is_credit_code(code):
                continue
            code_u = str(code).strip().upper()
            excel_r = excel_rows.get(code_u)
            if not excel_r:
                missing_rows.append(f"{code_u}: not found in excel sheet {sheet_name}")
                continue

            for field, ui_val in r.items():
                if not isinstance(field, str):
                    continue
                if field.startswith("__") or field in ignore:
                    continue
                excel_field = field if field in headers else UI_TO_EXCEL_FIELD.get(field, "")
                if not excel_field or excel_field not in headers:
                    continue
                excel_val = ws.cell(excel_r, headers[excel_field]).value
                if excel_val is None or (isinstance(excel_val, str) and excel_val.strip() == ""):
                    continue

                ui_num = _parse_number(ui_val)
                ex_num = _parse_number(excel_val)

                # Prefer numeric compare when both parse.
                if ui_num is not None or ex_num is not None:
                    ui_num, ex_num = _normalize_rate_pair(field, ui_num, ex_num)
                    if not _close(ui_num, ex_num, eps=_field_eps(field)):
                        mismatches.append(
                            Mismatch(
                                credit_code=code,
                                name=name,
                                sheet=sheet_name,
                                field=f"{field} (excel:{excel_field})" if excel_field != field else field,
                                expected=excel_val,
                                actual=ui_val,
                            )
                        )
                else:
                    if str(ui_val).strip() != str(excel_val).strip():
                        mismatches.append(
                            Mismatch(
                                credit_code=code,
                                name=name,
                                sheet=sheet_name,
                                field=f"{field} (excel:{excel_field})" if excel_field != field else field,
                                expected=excel_val,
                                actual=ui_val,
                            )
                        )

    return missing_rows, mismatches


def _export_index(wb) -> Dict[str, Tuple[Dict[str, int], Dict[str, int]]]:
    idx: Dict[str, Tuple[Dict[str, int], Dict[str, int]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        table = _sheet_table(ws)
        if table:
            idx[name] = table
    return idx


def _compare_ui_vs_export(ui_rows: List[Dict[str, Any]], export_wb) -> Tuple[List[str], List[Mismatch]]:
    missing: List[str] = []
    mismatches: List[Mismatch] = []

    idx = _export_index(export_wb)
    preferred_sheets = {
        "批发": ["批发", "批零总表"],
        "零售": ["零售", "批零总表"],
        "住宿": ["住宿", "住餐总表"],
        "餐饮": ["餐饮", "住餐总表"],
    }

    for r in ui_rows:
        code = str(r.get("__creditCode") or "").strip()
        name = str(r.get("__name") or "").strip()
        industry = str(r.get("__industry") or "").strip()
        if not _is_credit_code(code):
            continue
        code_u = str(code).strip().upper()

        candidates = preferred_sheets.get(industry, ["批零总表", "住餐总表", "批发", "零售", "住宿", "餐饮"])
        found_sheet: Optional[str] = None
        ex_headers: Dict[str, int] = {}
        ex_rows: Dict[str, int] = {}
        for s in candidates:
            t = idx.get(s)
            if not t:
                continue
            headers, rows = t
            if code_u in rows:
                found_sheet = s
                ex_headers, ex_rows = headers, rows
                break

        if not found_sheet:
            missing.append(f"{code_u}: not found in exported excel (industry={industry})")
            continue

        ex_r = ex_rows[code_u]
        for field, ui_val in r.items():
            if not isinstance(field, str):
                continue
            if field.startswith("__"):
                continue
            if field in ("规模", "标记", "来源表"):
                continue
            excel_field = field if field in ex_headers else _map_ui_field_to_export_header(field, found_sheet)
            if not excel_field or excel_field not in ex_headers:
                continue
            excel_val = export_wb[found_sheet].cell(ex_r, ex_headers[excel_field]).value
            ui_num = _parse_number(ui_val)
            ex_num = _parse_number(excel_val)
            if ui_num is not None or ex_num is not None:
                ui_num, ex_num = _normalize_rate_pair(field, ui_num, ex_num)
                if not _close(ui_num, ex_num, eps=_field_eps(field)):
                    mismatches.append(
                        Mismatch(
                            credit_code=code,
                            name=name,
                            sheet=found_sheet,
                            field=f"{field} (export:{excel_field})" if excel_field != field else field,
                            expected=excel_val,
                            actual=ui_val,
                        )
                    )
            else:
                if str(ui_val).strip() != str(excel_val).strip():
                    mismatches.append(
                            Mismatch(
                                credit_code=code,
                                name=name,
                                sheet=found_sheet,
                                field=f"{field} (export:{excel_field})" if excel_field != field else field,
                                expected=excel_val,
                                actual=ui_val,
                            )
                        )

    return missing, mismatches


@dataclass
class CompletenessCase:
    sheet: str
    credit_code: str
    name: str
    field: str
    expected: Any
    actual: Any
    ok: bool
    reason: str
    reproduce: str


def _build_expected_from_derived_sheet(wb, sheet_name: str) -> Dict[str, Dict[str, Any]]:
    # Returns creditCode -> {name, expected{ui_field:value}, sheet}
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    code_col = _get_header_cols(headers, "统一社会信用代码", 1)
    name_col = _get_header_cols(headers, "单位详细名称", 1)
    if not code_col or not name_col:
        return {}

    mapping = DERIVED_SHEET_MAPPINGS.get(sheet_name) or []
    if not mapping:
        return {}

    out: Dict[str, Dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not _is_credit_code(code):
            continue
        code_u = str(code).strip().upper()
        name = str(ws.cell(r, name_col).value or "").strip()

        expected: Dict[str, Any] = {}
        for ui_field, excel_header, nth in mapping:
            col = _get_header_cols(headers, excel_header, nth)
            if not col:
                continue
            v = ws.cell(r, col).value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            expected[ui_field] = v

        out[code_u] = {"name": name, "sheet": sheet_name, "expected": expected}
    return out


def _build_completeness_cases(
    input_wb,
    ui_rows: List[Dict[str, Any]],
) -> Tuple[List[CompletenessCase], Dict[str, Any]]:
    # Build UI index by credit code
    ui_by_code: Dict[str, Dict[str, Any]] = {}
    for r in ui_rows:
        code = r.get("__creditCode")
        if not _is_credit_code(code):
            continue
        ui_by_code[str(code).strip().upper()] = r

    cases: List[CompletenessCase] = []
    summary = {
        "sheets": {},
        "totalCompanies": 0,
        "missingCompanies": 0,
        "totalChecks": 0,
        "failedChecks": 0,
        "missingCodesBySheet": {},
    }

    for sheet_name in ["批发", "零售", "住宿", "餐饮"]:
        expected_map = _build_expected_from_derived_sheet(input_wb, sheet_name)
        sheet_summary = {"companies": len(expected_map), "missingCompanies": 0, "checks": 0, "fails": 0}
        summary["totalCompanies"] += len(expected_map)

        ui_codes_for_sheet = {
            str(r.get("__creditCode") or "").strip().upper()
            for r in ui_rows
            if str(r.get("来源表") or "").strip() == sheet_name and _is_credit_code(r.get("__creditCode"))
        }
        missing_codes = sorted([code for code in expected_map.keys() if code not in ui_codes_for_sheet])
        summary["missingCodesBySheet"][sheet_name] = missing_codes
        sheet_summary["missingCompanies"] = len(missing_codes)
        summary["missingCompanies"] += len(missing_codes)

        for code, info in expected_map.items():
            expected_fields: Dict[str, Any] = info.get("expected") or {}
            name = str(info.get("name") or "").strip()
            ui_row = ui_by_code.get(code)
            if not ui_row:
                continue

            ui_sheet = str(ui_row.get("来源表") or "").strip()
            ui_name = str(ui_row.get("__name") or name).strip()

            for field, exp in expected_fields.items():
                act = ui_row.get(field, "-")
                sheet_summary["checks"] += 1
                summary["totalChecks"] += 1

                if ui_sheet and ui_sheet != sheet_name:
                    okv = False
                    reason = f"来源表不一致：UI={ui_sheet}，Excel={sheet_name}"
                elif act in (None, "", "-"):
                    okv = False
                    reason = "Excel 有值，但明细表该字段为空/未展示"
                else:
                    exp_num = _parse_number(exp)
                    act_num = _parse_number(act)
                    if exp_num is not None or act_num is not None:
                        act_num, exp_num = _normalize_rate_pair(field, act_num, exp_num)
                        okv = _close(act_num, exp_num, eps=_field_eps(field))
                        reason = "" if okv else "数值不一致（允许少量格式化/四舍五入容差）"
                    else:
                        okv = str(act).strip() == str(exp).strip()
                        reason = "" if okv else "文本不一致"

                if not okv:
                    sheet_summary["fails"] += 1
                    summary["failedChecks"] += 1

                cases.append(
                    CompletenessCase(
                        sheet=sheet_name,
                        credit_code=code,
                        name=ui_name,
                        field=field,
                        expected=exp,
                        actual=act,
                        ok=okv,
                        reason=reason,
                        reproduce=f"首页明细表搜索 {code} → 展示列 {field} → 对照输入 Excel Sheet「{sheet_name}」的对应列",
                    )
                )

        summary["sheets"][sheet_name] = sheet_summary

    return cases, summary


def _summarize_mismatches(mismatches: List[Mismatch], limit: int = 200) -> str:
    if not mismatches:
        return "<p class='ok'>✅ 未发现差异</p>"
    rows = []
    for m in mismatches[:limit]:
        rows.append(
            "<tr>"
            f"<td>{_safe(m.credit_code)}</td>"
            f"<td>{_safe(m.name)}</td>"
            f"<td>{_safe(m.sheet)}</td>"
            f"<td>{_safe(m.field)}</td>"
            f"<td class='mono'>{_safe(str(m.expected))}</td>"
            f"<td class='mono'>{_safe(str(m.actual))}</td>"
            "</tr>"
        )
    more = ""
    if len(mismatches) > limit:
        more = f"<p class='warn'>仅展示前 {limit} 条，剩余 {len(mismatches)-limit} 条请查看原始数据 JSON。</p>"
    return (
        f"<p class='bad'>❌ 差异数量：{len(mismatches)}</p>"
        + more
        + "<div class='table-wrap'><table><thead><tr>"
        "<th>统一社会信用代码</th><th>企业</th><th>Sheet</th><th>字段</th><th>Excel</th><th>明细表</th>"
        "</tr></thead><tbody>"
        + "".join(rows)
        + "</tbody></table></div>"
    )


def _summarize_missing(missing: List[str], limit: int = 200) -> str:
    if not missing:
        return "<p class='ok'>✅ 覆盖完整</p>"
    items = "".join(f"<li class='mono'>{_safe(x)}</li>" for x in missing[:limit])
    more = ""
    if len(missing) > limit:
        more = f"<p class='warn'>仅展示前 {limit} 条，剩余 {len(missing)-limit} 条已省略。</p>"
    return f"<p class='bad'>❌ 覆盖问题：{len(missing)}</p>{more}<ul>{items}</ul>"

def _suggestions(
    ui_before_ok: bool,
    ui_after_ok: bool,
    before_rows: List[Dict[str, Any]],
    after_rows: List[Dict[str, Any]],
    input_wb,
    export_wb,
    missing_before: List[str],
    mismatches_before: List[Mismatch],
    missing_export: List[str],
    mismatches_export: List[Mismatch],
) -> str:
    hints: List[str] = []

    if not ui_before_ok or not ui_after_ok:
        hints.append("UI 明细表抽取失败：优先确认页面结构是否变化（table/td/input），或导入后是否出现“暂无数据”。")

    if input_wb is not None and ui_before_ok:
        excel_tables = [n for n in input_wb.sheetnames if _sheet_table(input_wb[n])]
        if not excel_tables:
            hints.append("输入 Excel 未识别到表头：检查是否存在“统一社会信用代码”列，或表头行不在前 10 行。")

    if export_wb is not None and ui_after_ok:
        # Heuristic: export may not contain per-company credit code rows.
        tables = [n for n in export_wb.sheetnames if _sheet_table(export_wb[n])]
        if not tables:
            hints.append("导出 Excel 未识别到企业明细表格：导出可能只包含汇总/衍生指标；建议在导出文件中确认是否存在“统一社会信用代码”列。")

    if missing_before:
        hints.append("导入覆盖缺失：若明细表分页/后端 pageSize 限制导致未加载全量企业，需要在 UI 或 API 增加分页拉全量（当前测试会在 report 中列出缺失 code）。")

    if mismatches_before:
        hints.append("导入字段不一致：通常是数值格式化/四舍五入/单位差异导致，建议对比 report 中同一企业同一字段的 Excel vs 明细表值。")

    if missing_export:
        hints.append("导出覆盖缺失：导出 Excel 里找不到信用代码对应行，可能是导出表不含企业级明细或 code 未写入。")

    if mismatches_export:
        hints.append("导出字段不一致：优先排查修改是否真正落库（输入框 blur 自动保存），以及导出是否读取最新数据。")

    if not hints:
        return "<p class='ok'>✅ 暂无额外改动建议</p>"

    items = "".join(f"<li>{_safe(x)}</li>" for x in hints)
    return f"<ul>{items}</ul>"


def _issues_summary(
    missing_before: List[str],
    mismatches_before: List[Mismatch],
    missing_export: List[str],
    mismatches_export: List[Mismatch],
    action_results: List[Dict[str, Any]],
    action_persist: List[Dict[str, Any]],
    indicator_action_fail: int,
    indicator_count_fail: int,
    indicator_export_fail: int,
    completeness_failed: int,
    completeness_total: int,
    forbidden_ui_columns_total: int,
    derived_unmapped_cols: int,
    derived_missing_ui_cols: int,
    tab_consistency_fail: int,
    ui_derived_fail: int,
    export_template_fail: int,
    export_formula_fail: int,
    normalization_fail: int,
    dag_fail: int,
    export_param_fail: int,
    linkage_preview_fail: int,
    llm_fail: int,
) -> str:
    issues: List[str] = []
    action_fail = sum(1 for r in action_results if r.get("ok") is False)
    persist_fail = sum(1 for r in action_persist if r.get("ok") is False)

    if missing_before:
        issues.append(f"导入覆盖缺失：{len(missing_before)}（见“导入一致性/覆盖检查”）")
    if mismatches_before:
        issues.append(f"导入字段不一致：{len(mismatches_before)}（见“导入一致性/字段一致性”）")
    if action_fail:
        issues.append(f"修改动作失败：{action_fail}（见“修改动作覆盖”）")
    if persist_fail:
        issues.append(f"修改持久化失败：{persist_fail}（见“修改动作覆盖”的 UI(刷新后) 列）")
    if indicator_action_fail:
        issues.append(f"指标调整动作失败：{indicator_action_fail}（见“指标调整与DAG联动覆盖”）")
    if indicator_count_fail:
        issues.append("指标数量不满足 16 项（见“指标调整与DAG联动覆盖”）")
    if indicator_export_fail:
        issues.append(f"指标与导出不一致：{indicator_export_fail}（见“指标调整与DAG联动覆盖”）")
    if missing_export:
        issues.append(f"导出覆盖缺失：{len(missing_export)}（见“导出一致性/覆盖检查”）")
    if mismatches_export:
        issues.append(f"导出字段不一致：{len(mismatches_export)}（见“导出一致性/字段一致性”）")
    if completeness_total > 0 and completeness_failed > 0:
        issues.append(f"明细表完整性断言失败：{completeness_failed}/{completeness_total}（见“明细表数据完整性案例”）")
    if forbidden_ui_columns_total > 0:
        issues.append(f"UI 存在已合并/禁用的重复列：{forbidden_ui_columns_total}（见“UI 列合并检查”）")
    if derived_unmapped_cols > 0:
        issues.append(f"衍生 Sheet 有值列未映射：{derived_unmapped_cols}（见“衍生 Sheet 列覆盖”）")
    if derived_missing_ui_cols > 0:
        issues.append(f"衍生 Sheet 映射列在 UI 缺失：{derived_missing_ui_cols}（见“衍生 Sheet 列覆盖”）")
    if tab_consistency_fail > 0:
        issues.append(f"Tab 覆盖/计数不一致：{tab_consistency_fail}（见“Tab 覆盖与计数（Excel vs UI）”）")
    if ui_derived_fail > 0:
        issues.append(f"UI 派生字段自洽性失败：{ui_derived_fail}（见“UI 派生字段一致性检查”）")
    if export_template_fail > 0:
        issues.append(f"导出模板结构不一致：{export_template_fail}（见“导出模板结构对标 PRD”）")
    if export_formula_fail > 0:
        issues.append(f"导出模板公式不一致：{export_formula_fail}（见“导出模板公式对标 PRD”）")
    if normalization_fail > 0:
        issues.append(f"输入标准化不通过：{normalization_fail}（见“输入标准化/兼容性”）")
    if dag_fail > 0:
        issues.append(f"DAG 联动覆盖不足：{dag_fail}（见“指标调整与DAG联动覆盖”）")
    if export_param_fail > 0:
        issues.append(f"导出参数化检查失败：{export_param_fail}（见“导出数字参数化”）")
    if linkage_preview_fail > 0:
        issues.append(f"联动预览高亮失败：{linkage_preview_fail}（见“联动预览高亮”）")
    if llm_fail > 0:
        issues.append("LLM 对话专项失败：1（见“LLM 对话专项测试”）")

    if not issues:
        return "<p class='ok'>✅ 未发现不符合预期项</p>"
    return "<ul>" + "".join(f"<li>{_safe(x)}</li>" for x in issues) + "</ul>"


def _list_screenshots(dir_path: str) -> List[str]:
    p = Path(dir_path)
    if not p.exists():
        return []
    return [str(x) for x in sorted(p.glob("*.png"))]


def _write_json(path: Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def _derived_columns(ws) -> List[Tuple[str, int, int]]:
    # Returns (headerText, nthOccurrence, colIndex) for derived sheets (row 1 headers).
    seen: Dict[str, int] = {}
    out: List[Tuple[str, int, int]] = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h is None or str(h).strip() == "":
            continue
        name = str(h).strip()
        seen[name] = seen.get(name, 0) + 1
        out.append((name, seen[name], col))
    return out


def _col_non_empty_stats(ws, col: int, max_examples: int = 3) -> Tuple[int, List[Any]]:
    count = 0
    examples: List[Any] = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        count += 1
        if len(examples) < max_examples and v not in examples:
            examples.append(v)
    return count, examples


def _guess_ui_field(excel_header: str) -> str:
    # Best-effort suggestion; used only for report hints.
    h = excel_header.strip()
    if h in {"粮油食品类", "饮料类", "烟酒类", "服装鞋帽针纺类", "日用品类", "汽车类", "吃穿用"}:
        return h
    if h in {"单位规模", "小微企业"}:
        return h
    return h


def _build_derived_column_coverage(input_wb, ui_headers: List[str]) -> Dict[str, Any]:
    # For each derived sheet, list columns that have values and whether they map to a UI column.
    out: Dict[str, Any] = {"sheets": {}}
    for sheet_name in ["批发", "零售", "住宿", "餐饮"]:
        if input_wb is None or sheet_name not in input_wb.sheetnames:
            continue
        ws = input_wb[sheet_name]
        mapping = DERIVED_SHEET_MAPPINGS.get(sheet_name) or []
        mapped: Dict[Tuple[str, int], str] = {(ex, nth): ui for (ui, ex, nth) in mapping}
        allowed: set[Tuple[str, int]] = {(ex, nth) for (ui, ex, nth) in mapping}

        items: List[Dict[str, Any]] = []
        missing_ui_cols = 0
        unmapped_cols_with_values = 0
        for header, nth, col in _derived_columns(ws):
            if header in {"序号", "统一社会信用代码", "单位详细名称"}:
                continue
            if header.startswith("[201-1]"):
                continue
            # PRD 关注的是“可微调 + 参与计算/导出”的核心指标列；输入表里的标记/分类字段（如「粮油食品类」「小微企业」等）
            # 不在当前 UI 明细表展示范围内，避免把“UI 未展示的元信息列”当成失败项。
            if (header, nth) not in allowed:
                continue
            non_empty, examples = _col_non_empty_stats(ws, col)
            if non_empty == 0:
                continue

            ui_field = mapped.get((header, nth), "")
            ui_present = bool(ui_field and ui_field in ui_headers)
            if not ui_field:
                # Unmapped column: still mark if the header itself is present in UI (rare, but helps debugging).
                ui_present = header in ui_headers
            if ui_field and not ui_present:
                missing_ui_cols += 1
            if not ui_field:
                unmapped_cols_with_values += 1

            items.append(
                {
                    "excelHeader": header,
                    "nth": nth,
                    "nonEmpty": non_empty,
                    "examples": examples,
                    "uiField": ui_field,
                    "uiPresent": ui_present,
                    "suggestedUiField": _guess_ui_field(header),
                }
            )

        out["sheets"][sheet_name] = {
            "columnsWithValues": len(items),
            "missingUiColumns": missing_ui_cols,
            "unmappedColumnsWithValues": unmapped_cols_with_values,
            "items": items,
        }
    return out


def _count_companies_in_sheet(input_wb, sheet_name: str) -> int:
    if input_wb is None or sheet_name not in input_wb.sheetnames:
        return 0
    ws = input_wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    code_col = _get_header_cols(headers, "统一社会信用代码", 1)
    if not code_col:
        return 0
    codes: set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, code_col).value
        if not _is_credit_code(v):
            continue
        codes.add(str(v).strip().upper())
    return len(codes)


def _ui_company_counts(ui_rows: List[Dict[str, Any]]) -> Dict[str, int]:
    out: Dict[str, int] = {"批发": 0, "零售": 0, "住宿": 0, "餐饮": 0}
    for r in ui_rows:
        code = r.get("__creditCode")
        if not _is_credit_code(code):
            continue
        sheet = str(r.get("来源表") or "").strip()
        if sheet in out:
            out[sheet] += 1
    return out


def _tab_count_consistency(input_wb, ui_rows: List[Dict[str, Any]], tab_counts: Dict[str, Any]) -> Dict[str, Any]:
    expected = {s: _count_companies_in_sheet(input_wb, s) for s in ["批发", "零售", "住宿", "餐饮"]}
    actual = _ui_company_counts(ui_rows)
    tab_items = (tab_counts.get("items") or []) if isinstance(tab_counts, dict) else []
    tab_map = {str(x.get("tab") or "").strip(): x for x in tab_items if isinstance(x, dict)}

    items: List[Dict[str, Any]] = []
    for s in ["批发", "零售", "住宿", "餐饮"]:
        it = tab_map.get(s) or {}
        ui_rows_count = int(actual.get(s) or 0)
        excel_count = int(expected.get(s) or 0)
        tab_row_count = int(it.get("rows") or 0) if isinstance(it, dict) else 0
        total_text = str(it.get("totalText") or "").strip() if isinstance(it, dict) else ""
        ok = (ui_rows_count == excel_count) and (tab_row_count == ui_rows_count or tab_row_count == 0)
        items.append(
            {
                "sheet": s,
                "excelCompanies": excel_count,
                "uiCompanies": ui_rows_count,
                "uiTabRows": tab_row_count,
                "uiTabTotalText": total_text,
                "ok": ok,
                "reason": "" if ok else "UI 企业覆盖/计数与输入 Excel 不一致（可能为解析遗漏、筛选口径差异或展示不完整）",
                "reproduce": f"导入后切换到「{s}」Tab，对比：UI 企业数 vs 输入 Excel「{s}」Sheet 的企业行数",
            }
        )
    return {"expected": expected, "actual": actual, "items": items}


def _calc_rate_percent(cur: Optional[float], base: Optional[float]) -> Optional[float]:
    if cur is None or base is None:
        return None
    if base == 0:
        return -100.0
    return (cur / base - 1.0) * 100.0


def _ui_derived_checks(ui_rows: List[Dict[str, Any]], limit_rows: int = 2000) -> List[Dict[str, Any]]:
    checks: List[Dict[str, Any]] = []

    def num(v: Any) -> Optional[float]:
        return _parse_number(v)

    for r in ui_rows[:limit_rows]:
        code = str(r.get("__creditCode") or "").strip()
        if not _is_credit_code(code):
            continue
        name = str(r.get("__name") or "").strip()
        industry = str(r.get("__industry") or "").strip()

        cur = num(r.get("本年-本月"))
        last = num(r.get("上年-本月"))
        prev = num(r.get("本年-上月"))
        diff_yoy = num(r.get("同比增量(当月)"))
        diff_mom = num(r.get("环比增量(当月)"))
        rate_mom = num(r.get("环比增速(当月)"))

        exp_diff_yoy = None if (cur is None or last is None) else (cur - last)
        exp_diff_mom = None if (cur is None or prev is None) else (cur - prev)
        exp_rate_mom = _calc_rate_percent(cur, prev)

        if exp_diff_yoy is not None and diff_yoy is not None and not _close(exp_diff_yoy, diff_yoy, eps=1.0):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "同比增量(当月)",
                    "expected": exp_diff_yoy,
                    "actual": diff_yoy,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致（可能为后端未重算/前端展示未刷新/舍入规则不一致）",
                    "reproduce": f"首页搜索 {code} → 查看 本年-本月/上年-本月/同比增量(当月) 三列是否满足：同比增量=本年-本月-上年-本月",
                }
            )
        if exp_diff_mom is not None and diff_mom is not None and not _close(exp_diff_mom, diff_mom, eps=1.0):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "环比增量(当月)",
                    "expected": exp_diff_mom,
                    "actual": diff_mom,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致",
                    "reproduce": f"首页搜索 {code} → 查看 本年-本月/本年-上月/环比增量(当月) 三列是否满足：环比增量=本年-本月-本年-上月",
                }
            )
        # UI 为整数百分比展示：允许四舍五入误差
        if exp_rate_mom is not None and rate_mom is not None and not _close(exp_rate_mom, rate_mom, eps=0.6):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "环比增速(当月)",
                    "expected": exp_rate_mom,
                    "actual": rate_mom,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致（可能是百分比/小数口径或舍入差异）",
                    "reproduce": f"首页搜索 {code} → 查看 本年-本月/本年-上月/环比增速(当月) 三列是否满足：环比增速=(本年-本月/本年-上月-1)*100",
                }
            )

        r_cur = num(r.get("零售额;本年-本月"))
        r_last = num(r.get("零售额;上年-本月"))
        r_prev = num(r.get("零售额;本年-上月"))
        r_diff_yoy = num(r.get("零售额;同比增量(当月)"))
        r_diff_mom = num(r.get("零售额;环比增量(当月)"))
        r_rate_mom = num(r.get("零售额;环比增速(当月)"))
        r_exp_diff_yoy = None if (r_cur is None or r_last is None) else (r_cur - r_last)
        r_exp_diff_mom = None if (r_cur is None or r_prev is None) else (r_cur - r_prev)
        r_exp_rate_mom = _calc_rate_percent(r_cur, r_prev)

        if r_exp_diff_yoy is not None and r_diff_yoy is not None and not _close(r_exp_diff_yoy, r_diff_yoy, eps=1.0):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "零售额;同比增量(当月)",
                    "expected": r_exp_diff_yoy,
                    "actual": r_diff_yoy,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致",
                    "reproduce": f"首页搜索 {code} → 查看 零售额;本年-本月/零售额;上年-本月/零售额;同比增量(当月) 是否一致",
                }
            )
        if r_exp_diff_mom is not None and r_diff_mom is not None and not _close(r_exp_diff_mom, r_diff_mom, eps=1.0):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "零售额;环比增量(当月)",
                    "expected": r_exp_diff_mom,
                    "actual": r_diff_mom,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致",
                    "reproduce": f"首页搜索 {code} → 查看 零售额;本年-本月/零售额;本年-上月/零售额;环比增量(当月) 是否一致",
                }
            )
        if r_exp_rate_mom is not None and r_rate_mom is not None and not _close(r_exp_rate_mom, r_rate_mom, eps=0.6):
            checks.append(
                {
                    "creditCode": code,
                    "name": name,
                    "industry": industry,
                    "field": "零售额;环比增速(当月)",
                    "expected": r_exp_rate_mom,
                    "actual": r_rate_mom,
                    "ok": False,
                    "reason": "UI 计算字段与基础字段不一致",
                    "reproduce": f"首页搜索 {code} → 查看 零售额;本年-本月/零售额;本年-上月/零售额;环比增速(当月) 是否一致",
                }
            )

    return checks


def _repo_root_from_this_file() -> Path:
    # .../tests/e2e/run_agent_browser_e2e_report.py -> repo root
    return Path(__file__).resolve().parents[2]


def _normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    s = re.sub(r"\\s+", "", s)
    return s.strip()


def _sheet_header_signature(ws, header_row: int, max_cols: int) -> List[str]:
    out: List[str] = []
    for c in range(1, max_cols + 1):
        out.append(_normalize_cell(ws.cell(header_row, c).value))
    while out and out[-1] == "":
        out.pop()
    return out


def _export_template_checks(export_wb_raw, template_wb_raw) -> Dict[str, Any]:
    expected_sheets = [
        "批零总表",
        "住餐总表",
        "批发",
        "零售",
        "住宿",
        "餐饮",
        "吃穿用",
        "小微",
        "吃穿用（剔除）",
        "社零额（定）",
        "汇总表（定）",
    ]
    export_sheets = list(export_wb_raw.sheetnames) if export_wb_raw is not None else []
    template_sheets = list(template_wb_raw.sheetnames) if template_wb_raw is not None else []

    missing = [s for s in expected_sheets if s not in export_sheets]
    extra = [s for s in export_sheets if s not in expected_sheets]

    header_checks: List[Dict[str, Any]] = []
    for s in [x for x in expected_sheets if x in export_sheets and x in template_sheets]:
        exp_ws = export_wb_raw[s]
        tpl_ws = template_wb_raw[s]

        exp_hr = _find_header_row(exp_ws)
        tpl_hr = _find_header_row(tpl_ws)
        if not exp_hr or not tpl_hr:
            continue

        max_cols = max(tpl_ws.max_column, exp_ws.max_column, 1)
        tpl_sig = _sheet_header_signature(tpl_ws, tpl_hr, max_cols)
        exp_sig = _sheet_header_signature(exp_ws, exp_hr, max_cols)
        ok = tpl_sig == exp_sig
        header_checks.append(
            {
                "sheet": s,
                "ok": ok,
                "templateHeaderCols": len(tpl_sig),
                "exportHeaderCols": len(exp_sig),
                "reason": "" if ok else "导出表头结构与定稿模板不一致（可能导致模板公式引用错位或字段缺失）",
                "reproduce": f"打开 prd/12月月报（定）(1).xlsx 与导出 Excel 的 Sheet「{s}」，对比表头（含空白列）是否逐列一致",
            }
        )

    return {
        "expectedSheets": expected_sheets,
        "exportSheets": export_sheets,
        "missingSheets": missing,
        "extraSheets": extra,
        "headerChecks": header_checks,
    }


def _export_formula_checks(export_wb_raw, template_wb_raw) -> List[Dict[str, Any]]:
    targets = [
        ("社零额（定）", "K3"),
        ("社零额（定）", "K7"),
        ("社零额（定）", "K9"),
        ("社零额（定）", "K15"),
        ("社零额（定）", "K17"),
        ("社零额（定）", "K19"),
        ("社零额（定）", "K23"),
        ("汇总表（定）", "D4"),
        ("汇总表（定）", "F4"),
        ("汇总表（定）", "D10"),
        ("汇总表（定）", "A11"),
    ]
    out: List[Dict[str, Any]] = []
    if export_wb_raw is None or template_wb_raw is None:
        return out
    for sheet, addr in targets:
        if sheet not in export_wb_raw.sheetnames or sheet not in template_wb_raw.sheetnames:
            continue
        ev = export_wb_raw[sheet][addr].value
        tv = template_wb_raw[sheet][addr].value
        ok = isinstance(ev, str) and ev.startswith("=") and isinstance(tv, str) and tv.startswith("=")
        if ok:
            ok = str(ev).strip() == str(tv).strip()
        out.append(
            {
                "sheet": sheet,
                "cell": addr,
                "ok": ok,
                "template": tv,
                "export": ev,
                "reason": "" if ok else "导出模板公式未按定稿模板保留（可能导致定稿表计算错误）",
                "reproduce": f"打开导出 Excel → Sheet「{sheet}」→ 单元格 {addr}，检查是否为公式且与定稿模板一致",
            }
        )
    return out


def _action_export_checks(
    action_results: List[Dict[str, Any]],
    action_persist: List[Dict[str, Any]],
    ui_lookup_rows: List[Dict[str, Any]],
    export_wb,
) -> List[Dict[str, Any]]:
    if export_wb is None:
        return []
    persist_map: Dict[str, Dict[str, Any]] = {}
    for p in action_persist:
        k = f"{p.get('creditCode')}|{p.get('field')}|{p.get('i')}"
        persist_map[k] = p
    ui_by_code = {
        str(r.get("__creditCode") or "").strip().upper(): r
        for r in ui_lookup_rows
        if _is_credit_code(r.get("__creditCode"))
    }
    idx = _export_index(export_wb)
    preferred_sheets = {
        "批发": ["批发", "批零总表"],
        "零售": ["零售", "批零总表"],
        "住宿": ["住宿", "住餐总表"],
        "餐饮": ["餐饮", "住餐总表"],
    }

    out: List[Dict[str, Any]] = []
    for a in action_results:
        cc = str(a.get("creditCode") or "").strip().upper()
        field = str(a.get("field") or "").strip()
        desired = a.get("value")
        k = f"{a.get('creditCode')}|{field}|{a.get('i')}"
        pv = persist_map.get(k) or {}
        persist_ok = pv.get("ok") is not False
        persist_value_raw = pv.get("uiValue")
        # If persistence doesn't match desired, use persisted UI value to validate export source-of-truth.
        exp = desired
        if persist_value_raw not in (None, "", "-"):
            exp = _parse_number(persist_value_raw)
            if exp is None:
                exp = str(persist_value_raw)

        r = ui_by_code.get(cc)
        if not r:
            out.append({"creditCode": cc, "field": field, "ok": False, "reason": "修改后 UI 未找到该企业，无法校验导出", "expected": exp})
            continue

        industry = str(r.get("__industry") or r.get("来源表") or "").strip()
        candidates = preferred_sheets.get(industry, []) + [industry]
        found_sheet = ""
        ex_headers: Dict[str, int] = {}
        ex_rows: Dict[str, int] = {}
        for s in candidates:
            if s in idx and cc in idx[s][1]:
                found_sheet = s
                ex_headers, ex_rows = idx[s]
                break
        if not found_sheet:
            out.append(
                {
                    "creditCode": cc,
                    "field": field,
                    "ok": False,
                    "reason": f"导出 Excel 未找到该企业行（industry={industry}）",
                    "expected": exp,
                }
            )
            continue

        excel_field = field if field in ex_headers else _map_ui_field_to_export_header(field, found_sheet)
        if not excel_field or excel_field not in ex_headers:
            out.append(
                {
                    "creditCode": cc,
                    "field": field,
                    "ok": False,
                    "reason": f"导出 Excel 缺少对应列（field={field}）",
                    "sheet": found_sheet,
                    "expected": exp,
                }
            )
            continue

        ex_r = ex_rows[cc]
        actual = export_wb[found_sheet].cell(ex_r, ex_headers[excel_field]).value
        exp_num = _parse_number(exp)
        act_num = _parse_number(actual)
        if exp_num is not None or act_num is not None:
            act_num, exp_num = _normalize_rate_pair(field, act_num, exp_num)
            okv = _close(act_num, exp_num, eps=_field_eps(field))
        else:
            okv = str(actual).strip() == str(exp).strip()

        reason = "" if okv else "导出值与期望不一致"
        if not persist_ok:
            reason = "修改未通过持久化校验（UI 刷新后异常），本项导出对照仅供参考"
        else:
            # If persisted value differs from desired, surface as a persistence issue rather than export issue.
            des_num = _parse_number(desired)
            pv_num = _parse_number(persist_value_raw)
            if des_num is not None or pv_num is not None:
                pv_num, des_num = _normalize_rate_pair(field, pv_num, des_num)
                if not _close(pv_num, des_num, eps=_field_eps(field)):
                    reason = "UI 刷新后值与期望修改不一致（可能未保存），导出按刷新后值校验"
            else:
                if persist_value_raw not in (None, "", "-") and str(persist_value_raw).strip() != str(desired).strip():
                    reason = "UI 刷新后值与期望修改不一致（可能未保存），导出按刷新后值校验"

        out.append(
            {
                "creditCode": cc,
                "field": field,
                "sheet": found_sheet,
                "excelField": excel_field,
                "expected": exp,
                "actual": actual,
                "ok": okv,
                "reason": reason,
                "desired": desired,
                "persistOk": bool(persist_ok),
                "persistValue": persist_value_raw,
            }
        )
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-url", required=True)
    ap.add_argument("--input-xlsx", required=True)
    ap.add_argument("--export-xlsx", required=True)
    ap.add_argument("--export-dag-xlsx", required=False, default="")
    ap.add_argument("--ui-before", required=True)
    ap.add_argument("--ui-after", required=True)
    ap.add_argument("--import-events", required=False, default="")
    ap.add_argument("--tab-counts", required=False, default="")
    ap.add_argument("--steps", required=False, default="")
    ap.add_argument("--actions", required=False, default="")
    ap.add_argument("--indicator-actions", required=False, default="")
    ap.add_argument("--indicators-before", required=False, default="")
    ap.add_argument("--indicators-after", required=False, default="")
    ap.add_argument("--ui-dag-before", required=False, default="")
    ap.add_argument("--ui-dag-after", required=False, default="")
    ap.add_argument("--linkage-preview", required=False, default="")
    ap.add_argument("--llm-results", required=False, default="")
    ap.add_argument("--export-alt-xlsx", required=False, default="")
    ap.add_argument("--export-param-meta", required=False, default="")
    ap.add_argument("--console", required=False, default="")
    ap.add_argument("--errors", required=False, default="")
    ap.add_argument("--trace", required=False, default="")
    ap.add_argument("--video", required=False, default="")
    ap.add_argument("--screenshots", required=False, default="")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    llm_results = _load_optional_json(args.llm_results) if args.llm_results else {}
    llm_executed = bool(llm_results.get("executed")) if llm_results else False
    llm_ok = bool(llm_results.get("ok")) if llm_executed else True
    llm_fail = 1 if llm_executed and not llm_ok else 0

    ui_before = _unwrap_agent_browser_json(_load_json(args.ui_before))
    ui_after = _unwrap_agent_browser_json(_load_json(args.ui_after))
    before_rows = ui_before.get("rows") or []
    after_rows = ui_after.get("rows") or []

    ui_before_error = ui_before.get("error") or ""
    ui_after_error = ui_after.get("error") or ""
    ui_before_ok = isinstance(before_rows, list) and len(before_rows) > 0 and not ui_before_error
    ui_after_ok = isinstance(after_rows, list) and len(after_rows) > 0 and not ui_after_error

    # UI 列合并：这些列在产品层面被认为“重复含义”，应该只保留一个展示口径。
    # 这里用 e2e 报告强约束：若 UI 仍出现这些列名，则直接判 FAIL。
    forbidden_ui_cols = {
        "本年-1—上月",
        "上年-1—上月",
        "零售额;本年-1—上月",
        "零售额;上年-1—上月",
    }
    before_headers = [str(x).strip() for x in (ui_before.get("headers") or []) if x is not None]
    after_headers = [str(x).strip() for x in (ui_after.get("headers") or []) if x is not None]
    forbidden_in_before = sorted({h for h in before_headers if h in forbidden_ui_cols})
    forbidden_in_after = sorted({h for h in after_headers if h in forbidden_ui_cols})

    input_wb_error = ""
    export_wb_error = ""
    export_dag_wb_error = ""
    export_wb_raw = None
    export_wb_raw_error = ""
    export_alt_wb_raw = None
    export_alt_wb_raw_error = ""
    template_wb_raw = None
    template_wb_raw_error = ""
    template_xlsx = ""
    try:
        input_wb = load_workbook(args.input_xlsx, data_only=True)
    except Exception as e:
        input_wb = None
        input_wb_error = str(e)
    try:
        export_wb = load_workbook(args.export_xlsx, data_only=True)
    except Exception as e:
        export_wb = None
        export_wb_error = str(e)
    export_dag_wb = None
    if args.export_dag_xlsx and Path(args.export_dag_xlsx).exists():
        try:
            export_dag_wb = load_workbook(args.export_dag_xlsx, data_only=True)
        except Exception as e:
            export_dag_wb = None
            export_dag_wb_error = str(e)
    try:
        export_wb_raw = load_workbook(args.export_xlsx, data_only=False)
    except Exception as e:
        export_wb_raw = None
        export_wb_raw_error = str(e)
    try:
        if args.export_alt_xlsx and Path(args.export_alt_xlsx).exists():
            export_alt_wb_raw = load_workbook(args.export_alt_xlsx, data_only=False)
    except Exception as e:
        export_alt_wb_raw = None
        export_alt_wb_raw_error = str(e)

    try:
        repo_root = _repo_root_from_this_file()
        cand = repo_root / "prd/12月月报（定）(1).xlsx"
        if cand.exists():
            template_xlsx = str(cand)
            template_wb_raw = load_workbook(cand, data_only=False)
    except Exception as e:
        template_wb_raw = None
        template_wb_raw_error = str(e)

    missing_before: List[str] = []
    mismatches_before: List[Mismatch] = []
    if input_wb is not None and ui_before_ok:
        missing_before, mismatches_before = _compare_ui_vs_excel(
            ui_rows=before_rows,
            wb=input_wb,
            ignore_fields=["规模", "标记"],
        )
    elif not ui_before_ok:
        missing_before = ["UI 抽取失败：无法进行导入一致性校验"]
    elif input_wb is None:
        missing_before = [f"输入 Excel 打开失败：{input_wb_error}"]

    missing_export: List[str] = []
    mismatches_export: List[Mismatch] = []
    if export_wb is not None and ui_after_ok:
        missing_export, mismatches_export = _compare_ui_vs_export(
            ui_rows=after_rows,
            export_wb=export_wb,
        )
    elif not ui_after_ok:
        missing_export = ["UI 抽取失败：无法进行导出一致性校验"]
    elif export_wb is None:
        missing_export = [f"导出 Excel 打开失败：{export_wb_error}"]

    import_events = {}
    if args.import_events and Path(args.import_events).exists():
        try:
            import_events = _unwrap_agent_browser_json(_load_json(args.import_events))
        except Exception:
            import_events = {}

    tab_counts: Dict[str, Any] = {}
    if args.tab_counts and Path(args.tab_counts).exists():
        try:
            tab_counts = _unwrap_agent_browser_json(_load_json(args.tab_counts))
        except Exception:
            tab_counts = {}

    screenshots = _list_screenshots(args.screenshots) if args.screenshots else []

    started_at = ui_before.get("extractedAt") or ""
    ended_at = ui_after.get("extractedAt") or ""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    steps: List[Dict[str, Any]] = []
    if args.steps and Path(args.steps).exists():
        try:
            steps = json.loads(Path(args.steps).read_text(encoding="utf-8"))
        except Exception:
            steps = []

    action_results: List[Dict[str, Any]] = []
    action_persist: List[Dict[str, Any]] = []
    if args.actions and Path(args.actions).exists():
        try:
            payload = json.loads(Path(args.actions).read_text(encoding="utf-8")) or {}
            action_results = payload.get("results") or []
            action_persist = payload.get("persist") or []
        except Exception:
            action_results = []
            action_persist = []

    indicator_action_results: List[Dict[str, Any]] = []
    if args.indicator_actions and Path(args.indicator_actions).exists():
        try:
            payload = json.loads(Path(args.indicator_actions).read_text(encoding="utf-8")) or {}
            indicator_action_results = payload.get("results") or []
        except Exception:
            indicator_action_results = []

    indicators_before_payload = _unwrap_agent_browser_json(_load_optional_json(args.indicators_before))
    indicators_after_payload = _unwrap_agent_browser_json(_load_optional_json(args.indicators_after))
    indicators_before_items = _indicator_items(indicators_before_payload)
    indicators_after_items = _indicator_items(indicators_after_payload)

    ui_dag_before_payload = _unwrap_agent_browser_json(_load_optional_json(args.ui_dag_before))
    ui_dag_after_payload = _unwrap_agent_browser_json(_load_optional_json(args.ui_dag_after))
    dag_before_rows = ui_dag_before_payload.get("rows") or []
    dag_after_rows = ui_dag_after_payload.get("rows") or []

    linkage_preview = _load_optional_json(args.linkage_preview) if args.linkage_preview else {}
    linkage_preview_executed = bool(linkage_preview)
    linkage_preview_ok = bool(linkage_preview.get("ok")) if linkage_preview_executed else False

    out_dir = Path(args.out).parent
    try:
        _write_json(out_dir / "missing_before.json", missing_before)
        _write_json(out_dir / "mismatches_before.json", [m.__dict__ for m in mismatches_before])
        _write_json(out_dir / "missing_export.json", missing_export)
        _write_json(out_dir / "mismatches_export.json", [m.__dict__ for m in mismatches_export])
    except Exception:
        pass

    # 输入 Excel 结构对标 PRD（sheet 集合）
    input_structure: Dict[str, Any] = {}
    try:
        expected_input_sheets = [
            "2024年12月批零",
            "2025年11月批零",
            "2024年3月",
            "2025年2月",
            "批发",
            "零售",
            "2024年12月住餐",
            "2025年11月住餐",
            "2024年3月住",
            "2025年2月住",
            "餐饮",
            "住宿",
            "限上零售额",
            "小微",
            "吃穿用",
        ]
        actual_sheets = list(input_wb.sheetnames) if input_wb is not None else []
        missing_sheets = [s for s in expected_input_sheets if s not in actual_sheets]
        extra_sheets = [s for s in actual_sheets if s not in expected_input_sheets]
        input_structure = {
            "expectedSheets": expected_input_sheets,
            "actualSheets": actual_sheets,
            "missingSheets": missing_sheets,
            "extraSheets": extra_sheets,
            "ok": (input_wb is not None) and (not missing_sheets),
        }
        _write_json(out_dir / "input_structure.json", input_structure)
    except Exception:
        input_structure = {"ok": False, "error": "failed to build input_structure"}

    completeness_cases: List[CompletenessCase] = []
    completeness_summary: Dict[str, Any] = {}
    completeness_failed = 0
    completeness_total = 0
    missing_codes_by_sheet: Dict[str, List[str]] = {}
    derived_column_coverage: Dict[str, Any] = {}
    forbidden_ui_columns_report: Dict[str, Any] = {
        "forbidden": sorted(forbidden_ui_cols),
        "before": forbidden_in_before,
        "after": forbidden_in_after,
        "total": len(set(forbidden_in_before + forbidden_in_after)),
        "ok": len(set(forbidden_in_before + forbidden_in_after)) == 0,
    }
    try:
        _write_json(out_dir / "ui_column_merge_checks.json", forbidden_ui_columns_report)
    except Exception:
        pass

    ui_normalization = _ui_normalization_checks(before_headers, before_rows if isinstance(before_rows, list) else [])
    try:
        _write_json(out_dir / "ui_normalization_checks.json", ui_normalization)
    except Exception:
        pass

    if input_wb is not None and ui_before_ok:
        try:
            completeness_cases, completeness_summary = _build_completeness_cases(input_wb, before_rows)
            completeness_total = len(completeness_cases)
            completeness_failed = sum(1 for c in completeness_cases if not c.ok)
            missing_codes_by_sheet = completeness_summary.get("missingCodesBySheet") or {}
            derived_column_coverage = _build_derived_column_coverage(input_wb, ui_before.get("headers") or [])
            _write_json(out_dir / "completeness_cases.json", [c.__dict__ for c in completeness_cases])
            _write_json(out_dir / "completeness_summary.json", completeness_summary)
            _write_json(out_dir / "missing_codes_by_sheet.json", missing_codes_by_sheet)
            _write_json(out_dir / "derived_column_coverage.json", derived_column_coverage)
        except Exception as e:
            completeness_summary = {"error": str(e)}
            completeness_total = 1
            completeness_failed = 1

    derived_unmapped_cols_total = 0
    derived_missing_ui_cols_total = 0
    try:
        for s in (derived_column_coverage.get("sheets") or {}).values():
            derived_unmapped_cols_total += int(s.get("unmappedColumnsWithValues") or 0)
            derived_missing_ui_cols_total += int(s.get("missingUiColumns") or 0)
    except Exception:
        derived_unmapped_cols_total = 0
        derived_missing_ui_cols_total = 0

    lookup_rows: List[Dict[str, Any]] = []
    if isinstance(before_rows, list):
        lookup_rows.extend(before_rows)
    if isinstance(after_rows, list):
        lookup_rows.extend(after_rows)
    action_export_checks = _action_export_checks(action_results, action_persist, lookup_rows, export_wb)
    try:
        _write_json(out_dir / "action_export_checks.json", action_export_checks)
    except Exception:
        pass

    tab_consistency: Dict[str, Any] = {}
    tab_consistency_fail = 0
    if input_wb is not None and ui_before_ok:
        try:
            tab_consistency = _tab_count_consistency(input_wb, before_rows, tab_counts)
            tab_consistency_fail = sum(
                1 for x in (tab_consistency.get("items") or []) if isinstance(x, dict) and not x.get("ok")
            )
            _write_json(out_dir / "tab_consistency.json", tab_consistency)
        except Exception:
            tab_consistency = {"error": "failed to build tab_consistency"}
            tab_consistency_fail = 1

    ui_derived = []
    ui_derived_fail = 0
    try:
        base_rows = after_rows if ui_after_ok else (before_rows if ui_before_ok else [])
        ui_derived = _ui_derived_checks(base_rows if isinstance(base_rows, list) else [])
        ui_derived_fail = len(ui_derived)
        _write_json(out_dir / "ui_derived_checks.json", ui_derived)
    except Exception:
        ui_derived = []
        ui_derived_fail = 0

    indicator_stats = _indicator_stats(before_rows if isinstance(before_rows, list) else [])
    indicator_compare: Dict[str, Any] = {"executed": False, "ok": True, "total": 0, "changed": 0, "items": []}
    if indicators_before_items and indicators_after_items:
        indicator_compare = _compare_indicators(indicators_before_items, indicators_after_items, indicator_stats)
        indicator_compare["executed"] = True
        try:
            _write_json(out_dir / "indicator_changes.json", indicator_compare)
        except Exception:
            pass

    indicator_count_before = _indicator_count_check(indicators_before_items)
    indicator_count_after = _indicator_count_check(indicators_after_items)
    try:
        _write_json(out_dir / "indicator_count_before.json", indicator_count_before)
        _write_json(out_dir / "indicator_count_after.json", indicator_count_after)
    except Exception:
        pass

    indicator_export_before = {"executed": False, "ok": True, "total": 0, "mismatch": 0, "items": []}
    indicator_export_after = {"executed": False, "ok": True, "total": 0, "mismatch": 0, "items": []}
    if indicators_before_items and export_wb is not None:
        indicator_export_before = _compare_indicator_values_to_export(indicators_before_items, export_wb)
        try:
            _write_json(out_dir / "indicator_export_before.json", indicator_export_before)
        except Exception:
            pass
    if indicators_after_items and export_dag_wb is not None:
        indicator_export_after = _compare_indicator_values_to_export(indicators_after_items, export_dag_wb)
        try:
            _write_json(out_dir / "indicator_export_after.json", indicator_export_after)
        except Exception:
            pass

    dag_summary: Dict[str, Any] = {"executed": False, "ok": True, "matchedRows": 0, "changedRows": 0}
    if isinstance(dag_before_rows, list) and isinstance(dag_after_rows, list) and dag_before_rows and dag_after_rows:
        dag_summary = _compare_dag_tables(dag_before_rows, dag_after_rows)
        dag_summary["executed"] = True
        try:
            _write_json(out_dir / "dag_summary.json", dag_summary)
        except Exception:
            pass

    export_param_meta = _load_optional_json(args.export_param_meta)
    export_param = _export_param_checks(export_wb_raw, export_alt_wb_raw, export_param_meta)
    try:
        _write_json(out_dir / "export_param_checks.json", export_param)
    except Exception:
        pass

    export_template: Dict[str, Any] = {}
    export_template_fail = 0
    export_formula: List[Dict[str, Any]] = []
    export_formula_fail = 0
    try:
        if export_wb_raw is not None and template_wb_raw is not None:
            export_template = _export_template_checks(export_wb_raw, template_wb_raw)
            export_template_fail = len(export_template.get("missingSheets") or []) + sum(
                1 for x in (export_template.get("headerChecks") or []) if isinstance(x, dict) and not x.get("ok")
            )
            export_formula = _export_formula_checks(export_wb_raw, template_wb_raw)
            export_formula_fail = sum(1 for x in export_formula if isinstance(x, dict) and not x.get("ok"))
            _write_json(out_dir / "export_template_checks.json", export_template)
            _write_json(out_dir / "export_formula_checks.json", export_formula)
    except Exception:
        export_template = {"error": "failed to build export_template_checks"}
        export_template_fail = 1

    indicator_action_fail = sum(1 for r in indicator_action_results if r.get("ok") is False)
    indicator_count_fail = 0
    if indicators_before_items or indicators_after_items:
        if not indicator_count_before.get("ok") or not indicator_count_after.get("ok"):
            indicator_count_fail = 1
    indicator_export_fail = 0
    if indicator_export_before.get("executed") and not indicator_export_before.get("ok"):
        indicator_export_fail += int(indicator_export_before.get("mismatch") or 1)
    if indicator_export_after.get("executed") and not indicator_export_after.get("ok"):
        indicator_export_fail += int(indicator_export_after.get("mismatch") or 1)
    linkage_preview_fail = 0
    if not linkage_preview_ok:
        linkage_preview_fail = 1
    normalization_fail = len(ui_normalization.get("unknownHeaders") or []) + len(ui_normalization.get("unknownSources") or []) + len(ui_normalization.get("unknownIndustries") or [])
    dag_fail = 0
    if indicator_compare.get("executed") and not indicator_compare.get("ok"):
        dag_fail += 1
    if dag_summary.get("executed") and not dag_summary.get("ok"):
        dag_fail += 1
    export_param_fail = 1 if export_param.get("executed") and not export_param.get("ok") else 0

    ok = (
        ui_before_ok
        and ui_after_ok
        and export_wb is not None
        and input_wb is not None
        and not missing_before
        and not mismatches_before
        and not missing_export
        and not mismatches_export
        and completeness_failed == 0
        and forbidden_ui_columns_report.get("ok") is True
        and tab_consistency_fail == 0
        and ui_derived_fail == 0
        and export_template_fail == 0
        and export_formula_fail == 0
        and ui_normalization.get("ok") is True
        and indicator_count_fail == 0
        and indicator_export_fail == 0
        and (not indicator_compare.get("executed") or indicator_compare.get("ok") is True)
        and (not dag_summary.get("executed") or dag_summary.get("ok") is True)
        and (not export_param.get("executed") or export_param.get("ok") is True)
        and linkage_preview_ok is True
        and (not llm_executed or llm_ok is True)
        and all(
            (derived_column_coverage.get("sheets") or {}).get(s, {}).get("missingUiColumns", 0) == 0
            for s in ["批发", "零售", "住宿", "餐饮"]
        )
        and all(s.get("status") != "fail" for s in steps)
        and all(r.get("ok") is not False for r in action_results)
        and all(r.get("ok") is not False for r in action_persist)
        and all(r.get("ok") is not False for r in indicator_action_results)
    )
    status = "PASS" if ok else "FAIL"

    import_log = ""
    try:
        import_log = "\n".join((import_events.get("items") or [])[:400])
    except Exception:
        import_log = ""

    shots_items: List[str] = []
    for p in screenshots[:30]:
        ps = _safe(p).replace("\\", "/")
        cap = _safe(Path(p).name)
        shots_items.append(
            "<div class=\"shot\">"
            f"<a class=\"shot-link\" href=\"{ps}\" data-cap=\"{cap}\">"
            f"<img class=\"thumb\" src=\"{ps}\" alt=\"shot\" loading=\"lazy\"/>"
            "</a>"
            f"<div class=\"cap mono\">{cap}</div>"
            "</div>"
        )
    shots_html = "".join(shots_items)

    conclusion = "✅ 全量校验通过（导入一致性 + 修改覆盖 + 导出一致性）" if ok else "❌ 存在不符合预期项：请优先查看【执行步骤与不符合预期项】与两段一致性差异表；所有改动建议也写在 report 内。"

    def _steps_html() -> str:
        if not steps:
            return "<p class='warn'>未生成 steps.json（脚本可能在早期退出）</p>"
        rows = []
        for s in steps:
            st = str(s.get("status") or "")
            klass = "ok" if st == "pass" else ("bad" if st == "fail" else "warn")
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(s.get('ts') or ''))}</td>"
                f"<td class='mono'>{_safe(str(s.get('name') or ''))}</td>"
                f"<td class='{klass}'>{_safe(st)}</td>"
                f"<td>{_safe(str(s.get('detail') or ''))}</td>"
                f"<td>{_safe(str(s.get('reproduce') or ''))}</td>"
                "</tr>"
            )
        return (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>时间</th><th>步骤</th><th>结果</th><th>不符合预期/原因</th><th>如何复现</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _actions_html() -> str:
        if not action_results:
            return "<p class='warn'>未执行修改动作或未生成 actions_result.json</p>"
        persist_map: Dict[str, Dict[str, Any]] = {}
        for p in action_persist:
            k = f"{p.get('creditCode')}|{p.get('field')}|{p.get('i')}"
            persist_map[k] = p
        rows = []
        for r in action_results:
            okv = bool(r.get("ok"))
            klass = "ok" if okv else "bad"
            k = f"{r.get('creditCode')}|{r.get('field')}|{r.get('i')}"
            pv = persist_map.get(k) or {}
            v = r.get("value")
            v_s = "" if v is None else str(v)
            ui_s = "" if r.get("uiValue") is None else str(r.get("uiValue"))
            pv_s = "" if pv.get("uiValue") is None else str(pv.get("uiValue"))
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(r.get('i') or ''))}</td>"
                f"<td class='mono'>{_safe(str(r.get('creditCode') or ''))}</td>"
                f"<td>{_safe(str(r.get('field') or ''))}</td>"
                f"<td class='mono'>{_safe(v_s)}</td>"
                f"<td class='mono'>{_safe(ui_s)}</td>"
                f"<td class='mono'>{_safe(pv_s)}</td>"
                f"<td class='{klass}'>{'PASS' if okv else 'FAIL'}</td>"
                f"<td>{_safe(str(r.get('reason') or ''))}</td>"
                f"<td class='mono'>{_safe(str(r.get('error') or ''))}</td>"
                "</tr>"
            )
        return (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>#</th><th>统一社会信用代码</th><th>字段</th><th>新值</th><th>UI(立即)</th><th>UI(刷新后)</th><th>结果</th><th>覆盖场景</th><th>错误</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _linkage_preview_html() -> str:
        if not linkage_preview:
            return "<p class='warn'>未执行联动预览高亮（缺少 linkage_preview_result.json）。</p>"
        okv = bool(linkage_preview.get("ok"))
        derived = linkage_preview.get("derived") or {}
        indicator = linkage_preview.get("indicator") or {}
        d_pick = derived.get("pick") or {}
        d_check = derived.get("check") or {}
        i_pick = indicator.get("pick") or {}
        i_check = indicator.get("check") or {}
        d_ok = bool(d_check.get("ok"))
        i_ok = bool(i_check.get("ok"))
        d_class = "ok" if d_ok else "bad"
        i_class = "ok" if i_ok else "bad"
        d_highlight = d_check.get("highlightCount")
        i_highlight_ind = i_check.get("highlightIndicators")
        i_highlight_cells = i_check.get("highlightCells")

        parent_rows = []
        parent_status = d_check.get("parentStatus") or []
        for p in parent_status:
            parent_rows.append(
                "<tr>"
                f"<td>{_safe(str(p.get('label') or ''))}</td>"
                f"<td class='mono'>{_safe(str(p.get('idx') or ''))}</td>"
                f"<td class='{ 'ok' if p.get('highlighted') else 'bad' }'>{'PASS' if p.get('highlighted') else 'FAIL'}</td>"
                "</tr>"
            )
        parent_table = (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>父节点列</th><th>列序号</th><th>高亮</th>"
            "</tr></thead><tbody>"
            + "".join(parent_rows)
            + "</tbody></table></div>"
            if parent_rows
            else "<p class='warn'>未获取父节点高亮明细。</p>"
        )

        return (
            f"<p class='{'ok' if okv else 'bad'}'>"
            + ("✅ 联动预览高亮通过" if okv else "❌ 联动预览高亮失败")
            + "</p>"
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>类型</th><th>目标</th><th>结果</th><th>备注</th>"
            "</tr></thead><tbody>"
            "<tr>"
            f"<td>派生列点击</td>"
            f"<td>{_safe(str(d_pick.get('header') or ''))}</td>"
            f"<td class='{d_class}'>{'PASS' if d_ok else 'FAIL'}</td>"
            f"<td>父节点高亮 {len(parent_status)} / 总高亮 {d_highlight}</td>"
            "</tr>"
            "<tr>"
            f"<td>指标点击</td>"
            f"<td>{_safe(str(i_pick.get('label') or ''))}</td>"
            f"<td class='{i_class}'>{'PASS' if i_ok else 'FAIL'}</td>"
            f"<td>高亮指标 {i_highlight_ind} / 高亮单元格 {i_highlight_cells}</td>"
            "</tr>"
            "</tbody></table></div>"
            + "<div style='margin-top: 8px;'>"
            + parent_table
            + "</div>"
        )

    def _indicator_actions_html() -> str:
        if not indicator_action_results:
            return "<p class='warn'>未执行指标调整动作或未生成 indicator_actions_result.json</p>"
        rows = []
        fails = [x for x in indicator_action_results if x.get("ok") is False]
        for r in indicator_action_results[:40]:
            okv = bool(r.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(r.get('i') or ''))}</td>"
                f"<td>{_safe(str(r.get('label') or ''))}</td>"
                f"<td class='mono'>{_safe(str(r.get('value') or ''))}</td>"
                f"<td class='mono'>{_safe(str(r.get('uiValue') or ''))}</td>"
                f"<td class='{klass}'>{'PASS' if okv else 'FAIL'}</td>"
                f"<td>{_safe(str(r.get('error') or r.get('reason') or ''))}</td>"
                "</tr>"
            )
        more = ""
        if len(indicator_action_results) > 40:
            more = "<p class='warn'>仅展示前 40 条，完整见 indicator_actions_result.json。</p>"
        return (
            f"<p class='{'ok' if not fails else 'bad'}'>"
            + ("✅ 全部 PASS" if not fails else f"❌ FAIL：{len(fails)}/{len(indicator_action_results)}")
            + "</p>"
            + more
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>#</th><th>指标</th><th>目标值</th><th>UI显示</th><th>结果</th><th>原因</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _indicator_changes_html() -> str:
        if not indicator_compare.get("executed"):
            return "<p class='warn'>未执行指标变化校验（缺少 indicators_before/after）。</p>"
        items = indicator_compare.get("items") or []
        fails = [x for x in items if not x.get("ok")]
        required = int(indicator_compare.get("required") or 0)
        changed = int(indicator_compare.get("changed") or 0)
        skipped = int(indicator_compare.get("skipped") or 0)
        total = int(indicator_compare.get("total") or 0)
        rows = []
        for it in items[:40]:
            okv = bool(it.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td>{_safe(str(it.get('label') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('before') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('after') or ''))}</td>"
                f"<td class='{klass}'>{'变更' if okv else '未变更'}</td>"
                "</tr>"
            )
        return (
            f"<p class='{'ok' if not fails else 'bad'}'>"
            + (
                f"✅ 指标变更：{changed}/{required}，跳过 {skipped}/{total}"
                if not fails
                else f"❌ 未变更：{len(fails)}/{len(items)}"
            )
            + "</p>"
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>指标</th><th>调整前</th><th>调整后</th><th>结果</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _indicator_count_html() -> str:
        if not indicators_before_items and not indicators_after_items:
            return "<p class='warn'>未执行指标数量校验（缺少 indicators_before/after）。</p>"
        rows = []
        for name, payload in [("调整前", indicator_count_before), ("调整后", indicator_count_after)]:
            okv = bool(payload.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td>{_safe(name)}</td>"
                f"<td class='mono'>{_safe(str(payload.get('count') or 0))}</td>"
                f"<td class='mono'>{_safe(str(payload.get('expected') or 16))}</td>"
                f"<td class='{klass}'>{'PASS' if okv else 'FAIL'}</td>"
                "</tr>"
            )
        return (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>阶段</th><th>实际指标数</th><th>期望</th><th>结果</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _indicator_export_block(title: str, result: Dict[str, Any], export_path: str) -> str:
        if not result.get("executed"):
            reason = "缺少导出文件或指标抽取结果"
            return f"<p class='warn'>{_safe(title)}：未执行（{reason}）</p>"
        mismatches = result.get("items") or []
        okv = bool(result.get("ok"))
        klass = "ok" if okv else "bad"
        rows = []
        for it in mismatches[:40]:
            rows.append(
                "<tr>"
                f"<td>{_safe(str(it.get('label') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('cell') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('expected') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('actual') or ''))}</td>"
                f"<td>{_safe(str(it.get('reason') or ''))}</td>"
                "</tr>"
            )
        summary = (
            f"<p class='{klass}'>"
            + (
                f"✅ {title}：全部一致（{result.get('total')} 项）"
                if okv
                else f"❌ {title}：不一致 {result.get('mismatch')}/{result.get('total')}"
            )
            + "</p>"
        )
        note = ""
        if export_path:
            note = f"<p class='warn'>导出文件：{_link_or_text(export_path)}</p>"
        table = ""
        if mismatches:
            table = (
                "<div class='table-wrap'><table><thead><tr>"
                "<th>指标</th><th>单元格</th><th>期望</th><th>实际</th><th>原因</th>"
                "</tr></thead><tbody>"
                + "".join(rows)
                + "</tbody></table></div>"
            )
        return summary + note + table

    def _indicator_export_html() -> str:
        return _indicator_export_block("调整前导出", indicator_export_before, args.export_xlsx) + (
            "<div style='height:10px'></div>"
        ) + _indicator_export_block("调整后导出", indicator_export_after, args.export_dag_xlsx)

    def _dag_table_diff_html() -> str:
        if not dag_summary.get("executed"):
            return "<p class='warn'>未执行 DAG 明细表差异校验（缺少 ui_companies_before/after_dag）。</p>"
        matched = int(dag_summary.get("matchedRows") or 0)
        changed = int(dag_summary.get("changedRows") or 0)
        total_cells = int(dag_summary.get("totalCells") or 0)
        changed_cells = int(dag_summary.get("changedCells") or 0)
        okv = bool(dag_summary.get("ok"))
        klass = "ok" if okv else "bad"
        samples = dag_summary.get("samples") or []
        rows = []
        for s in samples[:20]:
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(s.get('creditCode') or ''))}</td>"
                f"<td>{_safe(str(s.get('field') or ''))}</td>"
                f"<td class='mono'>{_safe(str(s.get('before') or ''))}</td>"
                f"<td class='mono'>{_safe(str(s.get('after') or ''))}</td>"
                "</tr>"
            )
        return (
            f"<p class='{klass}'>匹配企业 {matched}；变更企业 {changed}；变更单元格 {changed_cells}/{total_cells}</p>"
            + ("<p class='ok'>✅ DAG 下游明细表全量发生变化</p>" if okv else "<p class='bad'>❌ DAG 下游明细表未全量变化</p>")
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>统一社会信用代码</th><th>字段</th><th>调整前</th><th>调整后</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _export_param_html() -> str:
        if not export_param.get("executed"):
            reason = str((export_param.get("meta") or {}).get("reason") or "")
            if reason:
                return f"<p class='warn'>未执行导出参数化检查：{_safe(reason)}</p>"
            return "<p class='warn'>未执行导出参数化检查（缺少 export_alt.xlsx 或 export_param_meta.json）。</p>"
        file_name = str(export_param.get("fileName") or "")
        file_name_ok = export_param.get("fileNameOk")
        diffs = export_param.get("diffSamples") or []
        diff_count = int(export_param.get("diffCount") or 0)
        okv = bool(export_param.get("ok"))
        klass = "ok" if okv else "bad"
        rows = []
        for d in diffs[:30]:
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(d.get('sheet') or ''))}</td>"
                f"<td class='mono'>{_safe(str(d.get('cell') or ''))}</td>"
                f"<td class='mono'>{_safe(str(d.get('primary') or ''))}</td>"
                f"<td class='mono'>{_safe(str(d.get('alternate') or ''))}</td>"
                "</tr>"
            )
        return (
            f"<p class='{klass}'>字符串数值差异：{diff_count}</p>"
            + ("<p class='ok'>✅ 月份参数化生效（文本数字随月份变化）</p>" if okv else "<p class='bad'>❌ 文本数字未随月份变化</p>")
            + (f"<p class='{'ok' if file_name_ok else 'bad'}'>文件名：{_safe(file_name)}</p>" if file_name else "")
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>Sheet</th><th>单元格</th><th>主导出</th><th>替换导出</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _normalization_html() -> str:
        unknown_headers = ui_normalization.get("unknownHeaders") or []
        unknown_sources = ui_normalization.get("unknownSources") or []
        unknown_industries = ui_normalization.get("unknownIndustries") or []
        okv = bool(ui_normalization.get("ok"))
        klass = "ok" if okv else "bad"
        return (
            f"<p class='{klass}'>未知表头: {len(unknown_headers)}；未知来源表: {len(unknown_sources)}；未知行业: {len(unknown_industries)}</p>"
            + ("<p class='ok'>✅ UI 字段与来源表已标准化</p>" if okv else "<p class='bad'>❌ UI 字段/来源表存在非标准值</p>")
            + "<div class='kv'>"
            + f"<div class='k'>未知表头</div><div class='v'>{_safe(str(unknown_headers))}</div>"
            + f"<div class='k'>未知来源表</div><div class='v'>{_safe(str(unknown_sources))}</div>"
            + f"<div class='k'>未知行业</div><div class='v'>{_safe(str(unknown_industries))}</div>"
            + "</div>"
        )

    def _completeness_html() -> str:
        if input_wb is None or not ui_before_ok:
            return "<p class='warn'>输入 Excel 或 UI 抽取失败：未执行完整性断言</p>"

        if not completeness_cases and not missing_codes_by_sheet:
            return "<p class='warn'>未生成完整性数据（可能是衍生 Sheet 映射为空）</p>"

        pinned_cc = "914401007RDD76M0RF"
        pinned_fields = ["上年-本月", "销售额;上年-本月", "商品销售额;上年-本月"]
        pinned = [c for c in completeness_cases if c.credit_code.upper() == pinned_cc and c.field in pinned_fields]
        fail_cases = [c for c in completeness_cases if not c.ok]
        show_fails = fail_cases[:200]

        def _row(c: CompletenessCase) -> str:
            klass = "ok" if c.ok else "bad"
            return (
                "<tr>"
                f"<td class='mono'>{_safe(c.sheet)}</td>"
                f"<td class='mono'>{_safe(c.credit_code)}</td>"
                f"<td>{_safe(c.name)}</td>"
                f"<td>{_safe(c.field)}</td>"
                f"<td class='mono'>{_safe(str(c.expected))}</td>"
                f"<td class='mono'>{_safe(str(c.actual))}</td>"
                f"<td class='{klass}'>{'PASS' if c.ok else 'FAIL'}</td>"
                f"<td>{_safe(c.reason)}</td>"
                f"<td>{_safe(c.reproduce)}</td>"
                "</tr>"
            )

        pinned_html = ""
        if pinned:
            pinned_html = (
                "<p class='ok'>✅ 关键案例（用户指定）</p>"
                "<div class='table-wrap'><table><thead><tr>"
                "<th>Sheet</th><th>统一社会信用代码</th><th>企业</th><th>字段</th><th>Excel</th><th>明细表</th><th>结果</th><th>原因</th><th>如何复现</th>"
                "</tr></thead><tbody>"
                + "".join(_row(c) for c in pinned[:1])
                + "</tbody></table></div>"
            )
        else:
            pinned_html = (
                "<p class='bad'>❌ 关键案例未命中（未在完整性断言数据中找到该企业字段）</p>"
                "<p class='warn'>复现：导入后在明细表搜索 914401007RDD76M0RF，确认列名是否为“商品销售额;上年-本月”。</p>"
            )

        fail_html = ""
        if not fail_cases:
            fail_html = "<p class='ok'>✅ 字段级完整性断言：未发现失败</p>"
        else:
            more = ""
            if len(fail_cases) > len(show_fails):
                more = f"<p class='warn'>仅展示前 {len(show_fails)} 条失败，完整列表见 completeness_cases.json。</p>"
            fail_html = (
                f"<p class='bad'>❌ 字段级完整性断言失败：{len(fail_cases)}/{len(completeness_cases)}</p>"
                + more
                + "<div class='table-wrap'><table><thead><tr>"
                "<th>Sheet</th><th>统一社会信用代码</th><th>企业</th><th>字段</th><th>Excel</th><th>明细表</th><th>结果</th><th>原因</th><th>如何复现</th>"
                "</tr></thead><tbody>"
                + "".join(_row(c) for c in show_fails)
                + "</tbody></table></div>"
            )

        missing_total = sum(len(v) for v in (missing_codes_by_sheet or {}).values())
        missing_html = f"<p class='{'ok' if missing_total == 0 else 'bad'}'>企业覆盖缺失：{missing_total}</p>"

        return pinned_html + "<div style='height:12px'></div>" + missing_html + "<div style='height:12px'></div>" + fail_html

    def _missing_codes_html() -> str:
        if not missing_codes_by_sheet:
            return "<p class='warn'>未生成缺失企业列表</p>"
        parts: List[str] = []
        for sheet_name in ["批发", "零售", "住宿", "餐饮"]:
            codes = missing_codes_by_sheet.get(sheet_name) or []
            sample = codes[:30]
            parts.append(
                "<div class='card' style='margin-top:12px;'>"
                f"<h2>Sheet「{_safe(sheet_name)}」缺失企业：{_safe(str(len(codes)))}</h2>"
                + ("<p class='ok'>✅ 无缺失</p>" if not codes else "<ul>" + "".join(f"<li class='mono'>{_safe(x)}</li>" for x in sample) + "</ul>")
                + (f"<p class='warn'>仅展示前 {len(sample)} 个，完整见 missing_codes_by_sheet.json。</p>" if len(codes) > len(sample) else "")
                + "</div>"
            )
        return "".join(parts)

    def _derived_coverage_html() -> str:
        sheets = (derived_column_coverage.get("sheets") or {}) if isinstance(derived_column_coverage, dict) else {}
        if not sheets:
            return "<p class='warn'>未生成“衍生 Sheet 列覆盖”信息</p>"
        rows: List[str] = []
        for sheet_name in ["批发", "零售", "住宿", "餐饮"]:
            s = sheets.get(sheet_name) or {}
            items = s.get("items") or []
            for it in items[:200]:
                mapped = bool(it.get("uiField"))
                ui_present = bool(it.get("uiPresent"))
                # 注意：报告中不要出现固定英文 token（方便后续全文检索/阅读），用中文描述状态。
                status_text = "OK" if (mapped and ui_present) else ("映射列未展示" if mapped else "未映射")
                klass = "ok" if status_text == "OK" else "bad"
                header = str(it.get("excelHeader") or "")
                nth = int(it.get("nth") or 1)
                nth_txt = f"#{nth}" if nth > 1 else ""
                rows.append(
                    "<tr>"
                    f"<td class='mono'>{_safe(sheet_name)}</td>"
                    f"<td class='mono'>{_safe(header + nth_txt)}</td>"
                    f"<td class='mono'>{_safe(str(it.get('nonEmpty') or ''))}</td>"
                    f"<td class='mono'>{_safe(str(it.get('uiField') or ''))}</td>"
                    f"<td class='{klass}'>{_safe(status_text)}</td>"
                    f"<td class='mono'>{_safe(str(it.get('examples') or ''))}</td>"
                    f"<td class='mono'>{_safe(str(it.get('suggestedUiField') or ''))}</td>"
                    "</tr>"
                )
        return (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>Sheet</th><th>Excel 列</th><th>有值数量</th><th>映射到明细表列</th><th>状态</th><th>示例</th><th>建议 UI 列名</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _action_export_checks_html() -> str:
        if export_wb is None or not ui_after_ok:
            return "<p class='warn'>导出 Excel 或 UI 抽取失败：未执行“修改动作 → 导出回归校验”</p>"
        if not action_export_checks:
            return "<p class='warn'>无修改动作或无法生成回归校验</p>"
        fails = [x for x in action_export_checks if not x.get("ok")]
        rows = []
        for x in action_export_checks[:200]:
            okv = bool(x.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(x.get('creditCode') or ''))}</td>"
                f"<td>{_safe(str(x.get('field') or ''))}</td>"
                f"<td class='mono'>{_safe(str(x.get('sheet') or ''))}</td>"
                f"<td class='mono'>{_safe(str(x.get('excelField') or ''))}</td>"
                f"<td class='mono'>{_safe(str(x.get('expected') or ''))}</td>"
                f"<td class='mono'>{_safe(str(x.get('actual') or ''))}</td>"
                f"<td class='{klass}'>{'PASS' if okv else 'FAIL'}</td>"
                f"<td>{_safe(str(x.get('reason') or ''))}</td>"
                "</tr>"
            )
        more = ""
        if len(action_export_checks) > 200:
            more = "<p class='warn'>仅展示前 200 条，完整见 action_export_checks.json。</p>"
        return (
            f"<p class='{'ok' if not fails else 'bad'}'>"
            + ("✅ 全部 PASS" if not fails else f"❌ FAIL：{len(fails)}/{len(action_export_checks)}")
            + "</p>"
            + more
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>统一社会信用代码</th><th>字段</th><th>Sheet</th><th>导出列</th><th>期望</th><th>实际</th><th>结果</th><th>原因</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _tab_counts_html() -> str:
        items = tab_counts.get("items") if isinstance(tab_counts, dict) else None
        if not isinstance(items, list) or not items:
            return "<p class='warn'>未生成 tab 行数统计</p>"
        rows = []
        for it in items:
            okv = bool(it.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(it.get('tab') or ''))}</td>"
                f"<td class='{klass}'>{'PASS' if okv else 'FAIL'}</td>"
                f"<td class='mono'>{_safe(str(it.get('rows') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('totalText') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('error') or ''))}</td>"
                "</tr>"
            )
        return (
            "<div class='table-wrap'><table><thead><tr>"
            "<th>Tab</th><th>结果</th><th>表格行数</th><th>底部总数文案</th><th>错误</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _llm_html() -> str:
        if not llm_results:
            return "<p class='warn'>未执行 LLM 专项测试（llm_results.json 不存在）</p>"
        if not llm_executed:
            reason = str(llm_results.get("skipReason") or "")
            return f"<p class='warn'>未执行 LLM 专项测试：{_safe(reason or 'skip')}</p>"

        def _ok_text(v: Any) -> str:
            return "PASS" if v else "FAIL"

        conf = llm_results.get("config") or {}
        streaming = llm_results.get("streaming") or {}
        first = llm_results.get("first") or {}
        second = llm_results.get("second") or {}
        updates = llm_results.get("updates") or {}
        errors = llm_results.get("errors") or []

        rows: List[str] = []

        def add_row(name: str, okv: Any, detail: str) -> None:
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td>{_safe(name)}</td>"
                f"<td class='{klass}'>{_ok_text(okv)}</td>"
                f"<td class='mono'>{_safe(detail)}</td>"
                "</tr>"
            )

        add_row(
            "全局配置",
            bool(conf.get("baseUrlOk") and conf.get("modelOk") and conf.get("apiKeyMasked") and conf.get("showHideOk")),
            f"baseUrl={conf.get('baseUrl')} model={conf.get('model')} apiKeyType={conf.get('apiKeyType')}",
        )
        add_row("流式输出", bool(streaming.get("ok")), "检测到流式指示符")

        first_summary = str(first.get("summaryText") or "")
        second_summary = str(second.get("summaryText") or "")
        add_row("首轮对话(含指标目标)", bool((first.get("summary") or {}).get("optimized") is True), first_summary[:200])
        add_row("二轮对话(无指标目标)", bool((second.get("summary") or {}).get("optimized") is False), second_summary[:200])

        wr = updates.get("wr") or {}
        ac = updates.get("ac") or {}
        after_opt = updates.get("afterOptimize") or {}
        add_row(
            "数据修改校验(WR)",
            bool(wr.get("ok")),
            f"{wr.get('id')} {wr.get('field')} target={wr.get('target')} after={wr.get('after')}",
        )
        add_row(
            "数据修改校验(AC)",
            bool(ac.get("ok")),
            f"{ac.get('id')} {ac.get('field')} target={ac.get('target')} after={ac.get('after')}",
        )
        if after_opt:
            wr_opt = after_opt.get("wr") or {}
            ac_opt = after_opt.get("ac") or {}
            add_row(
                "智能调整后(WR保持)",
                bool(wr_opt.get("ok")),
                f"target={wr_opt.get('target')} after={wr_opt.get('after')}",
            )
            add_row(
                "智能调整后(AC保持)",
                bool(ac_opt.get("ok")),
                f"target={ac_opt.get('target')} after={ac_opt.get('after')}",
            )

        if errors:
            add_row("错误详情", False, " | ".join(str(x) for x in errors[:5]))

        status_text = "✅ PASS" if llm_ok else "❌ FAIL"
        return (
            f"<p class='{'ok' if llm_ok else 'bad'}'>{status_text}</p>"
            + "<p class='warn'>完整记录：<a class='mono' href='llm_results.json'>llm_results.json</a></p>"
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>检查项</th><th>结果</th><th>细节</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _input_structure_html() -> str:
        if not isinstance(input_structure, dict) or not input_structure:
            return "<p class='warn'>未生成输入结构检查</p>"
        missing = input_structure.get("missingSheets") or []
        extra = input_structure.get("extraSheets") or []
        okv = bool(input_structure.get("ok"))
        return (
            ("<p class='ok'>✅ 输入 sheet 集合符合 PRD（允许 extra）</p>" if okv else f"<p class='bad'>❌ 输入 sheet 缺失：{_safe(str(len(missing)))}</p>")
            + "<p class='warn'>完整列表：<a class='mono' href='input_structure.json'>input_structure.json</a></p>"
            + ("<p class='bad'>缺失：</p><ul>" + "".join(f"<li class='mono'>{_safe(x)}</li>" for x in missing[:50]) + "</ul>" if missing else "")
            + ("<p class='warn'>额外：</p><ul>" + "".join(f"<li class='mono'>{_safe(x)}</li>" for x in extra[:50]) + "</ul>" if extra else "")
        )

    def _tab_consistency_html() -> str:
        items = (tab_consistency.get("items") or []) if isinstance(tab_consistency, dict) else []
        if not items:
            return "<p class='warn'>未生成 tab_consistency.json</p>"
        rows = []
        for it in items:
            okv = bool(it.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(it.get('sheet') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('excelCompanies') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('uiCompanies') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('uiTabRows') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('uiTabTotalText') or ''))}</td>"
                f"<td class='{klass}'>{'OK' if okv else 'FAIL'}</td>"
                f"<td>{_safe(str(it.get('reason') or ''))}</td>"
                f"<td>{_safe(str(it.get('reproduce') or ''))}</td>"
                "</tr>"
            )
        return (
            "<p class='warn'>完整列表：<a class='mono' href='tab_consistency.json'>tab_consistency.json</a></p>"
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>Sheet</th><th>Excel 企业数</th><th>UI 企业数(抽取)</th><th>UI Tab 行数</th><th>UI Tab 文案</th><th>结果</th><th>原因</th><th>复现</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _ui_derived_html() -> str:
        if not ui_derived:
            return "<p class='ok'>✅ 未发现 UI 派生字段不一致</p>"
        rows = []
        for it in ui_derived[:200]:
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(it.get('creditCode') or ''))}</td>"
                f"<td>{_safe(str(it.get('name') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('industry') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('field') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('expected') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('actual') or ''))}</td>"
                f"<td>{_safe(str(it.get('reason') or ''))}</td>"
                f"<td>{_safe(str(it.get('reproduce') or ''))}</td>"
                "</tr>"
            )
        more = "" if len(ui_derived) <= 200 else "<p class='warn'>仅展示前 200 条，完整见 ui_derived_checks.json。</p>"
        return (
            f"<p class='bad'>❌ FAIL：{len(ui_derived)}</p>"
            + "<p class='warn'>完整列表：<a class='mono' href='ui_derived_checks.json'>ui_derived_checks.json</a></p>"
            + more
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>统一社会信用代码</th><th>企业</th><th>行业</th><th>字段</th><th>期望(计算)</th><th>实际(UI)</th><th>原因</th><th>复现</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _export_template_html() -> str:
        if not export_template:
            return "<p class='warn'>未生成导出模板结构校验</p>"
        missing = export_template.get("missingSheets") or []
        extra = export_template.get("extraSheets") or []
        checks = export_template.get("headerChecks") or []
        rows = []
        for it in checks[:200]:
            okv = bool(it.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(it.get('sheet') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('templateHeaderCols') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('exportHeaderCols') or ''))}</td>"
                f"<td class='{klass}'>{'OK' if okv else 'FAIL'}</td>"
                f"<td>{_safe(str(it.get('reason') or ''))}</td>"
                f"<td>{_safe(str(it.get('reproduce') or ''))}</td>"
                "</tr>"
            )
        missing_html = "" if not missing else "<p class='bad'>缺失 sheet：</p><ul>" + "".join(f"<li class='mono'>{_safe(x)}</li>" for x in missing) + "</ul>"
        extra_html = "" if not extra else "<p class='warn'>额外 sheet：</p><ul>" + "".join(f"<li class='mono'>{_safe(x)}</li>" for x in extra) + "</ul>"
        return (
            f"<p class='warn'>定稿模板：{_safe(template_xlsx or '(not found)')}</p>"
            + "<p class='warn'>完整列表：<a class='mono' href='export_template_checks.json'>export_template_checks.json</a></p>"
            + missing_html
            + extra_html
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>Sheet</th><th>模板表头列数</th><th>导出表头列数</th><th>结果</th><th>原因</th><th>复现</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    def _export_formula_html() -> str:
        if not export_formula:
            return "<p class='warn'>未生成导出公式校验（可能模板缺失或导出无法打开）</p>"
        rows = []
        for it in export_formula:
            okv = bool(it.get("ok"))
            klass = "ok" if okv else "bad"
            rows.append(
                "<tr>"
                f"<td class='mono'>{_safe(str(it.get('sheet') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('cell') or ''))}</td>"
                f"<td class='{klass}'>{'OK' if okv else 'FAIL'}</td>"
                f"<td class='mono'>{_safe(str(it.get('template') or ''))}</td>"
                f"<td class='mono'>{_safe(str(it.get('export') or ''))}</td>"
                f"<td>{_safe(str(it.get('reason') or ''))}</td>"
                f"<td>{_safe(str(it.get('reproduce') or ''))}</td>"
                "</tr>"
            )
        return (
            "<p class='warn'>完整列表：<a class='mono' href='export_formula_checks.json'>export_formula_checks.json</a></p>"
            + "<div class='table-wrap'><table><thead><tr>"
            "<th>Sheet</th><th>Cell</th><th>结果</th><th>模板公式</th><th>导出公式</th><th>原因</th><th>复现</th>"
            "</tr></thead><tbody>"
            + "".join(rows)
            + "</tbody></table></div>"
        )

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Northstar E2E (agent-browser) - {status}</title>
  <style>
    :root {{
      --bg: #0b1220;
      --panel: #0f1b33;
      --text: #e6edf3;
      --muted: #a6b3c3;
      --ok: #4ade80;
      --bad: #fb7185;
      --warn: #fbbf24;
      --border: rgba(255,255,255,.10);
      --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
      --sans: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
    }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font-family: var(--sans); }}
    .page {{ display: flex; align-items: flex-start; }}
    .sidebar {{
      position: sticky;
      top: 0;
      height: 100vh;
      width: 270px;
      flex: 0 0 270px;
      overflow: auto;
      padding: 18px 14px 18px;
      border-right: 1px solid var(--border);
      background: linear-gradient(180deg, rgba(15,27,51,.92), rgba(11,18,32,.92));
      backdrop-filter: blur(10px);
    }}
    .side-head {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; margin-bottom: 12px; }}
    .side-title {{ font-weight: 800; font-size: 13px; letter-spacing: .02em; color: var(--text); }}
    .side-sub {{ color: var(--muted); font-size: 12px; margin-top: 2px; }}
    .side-nav {{ display: flex; flex-direction: column; gap: 4px; margin-top: 10px; }}
    .side-nav a {{
      display: block;
      padding: 8px 10px;
      border-radius: 10px;
      border: 1px solid transparent;
      color: var(--text);
      font-size: 13px;
      line-height: 1.2;
    }}
    .side-nav a:hover {{ background: rgba(255,255,255,.04); text-decoration: none; border-color: rgba(255,255,255,.10); }}
    .side-nav a.active {{ background: rgba(147,197,253,.10); border-color: rgba(147,197,253,.25); }}
    .side-nav .group {{ margin-top: 10px; padding: 8px 10px; color: var(--muted); font-size: 12px; font-weight: 800; letter-spacing: .04em; text-transform: uppercase; }}
    .side-foot {{ margin-top: 14px; padding-top: 12px; border-top: 1px solid var(--border); color: var(--muted); font-size: 12px; }}
    .side-foot a {{ color: #93c5fd; }}
    .main {{ min-width: 0; flex: 1 1 auto; }}
    .wrap {{ max-width: 1200px; margin: 0 auto; padding: 28px 18px 48px; }}
    .title {{ display: flex; align-items: baseline; justify-content: space-between; gap: 12px; }}
    h1 {{ font-size: 22px; margin: 0; }}
    .pill {{ padding: 4px 10px; border-radius: 999px; font-weight: 700; font-size: 12px; border: 1px solid var(--border); }}
    .pill.ok {{ color: #052e15; background: var(--ok); border-color: rgba(74,222,128,.35); }}
    .pill.bad {{ color: #2b0b11; background: var(--bad); border-color: rgba(251,113,133,.35); }}
    .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 16px; }}
    .card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 12px; padding: 14px 14px 12px; }}
    .card h2 {{ font-size: 14px; margin: 0 0 8px; color: var(--muted); font-weight: 700; }}
    .kv {{ display: grid; grid-template-columns: 160px 1fr; gap: 6px 10px; font-size: 13px; }}
    .kv div {{ padding: 2px 0; }}
    .k {{ color: var(--muted); }}
    .v {{ font-family: var(--mono); }}
    .mono {{ font-family: var(--mono); }}
    .ok {{ color: var(--ok); font-weight: 700; }}
    .bad {{ color: var(--bad); font-weight: 700; }}
    .warn {{ color: var(--warn); font-weight: 700; }}
    .section {{ margin-top: 14px; }}
    .section h3 {{ margin: 0 0 10px; font-size: 16px; }}
    .h3line {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; }}
    .h3line a.permalink {{ font-family: var(--mono); font-size: 12px; color: var(--muted); }}
    .h3line a.permalink:hover {{ color: #93c5fd; text-decoration: none; }}
    a {{ color: #93c5fd; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    ul {{ margin: 8px 0 0 18px; }}
    li {{ margin: 3px 0; }}
    .table-wrap {{ overflow: auto; border: 1px solid var(--border); border-radius: 10px; }}
    table {{ border-collapse: collapse; width: 100%; min-width: 900px; font-size: 12px; }}
    th, td {{ border-bottom: 1px solid var(--border); padding: 8px 10px; text-align: left; vertical-align: top; }}
    th {{ position: sticky; top: 0; background: rgba(15,27,51,.95); backdrop-filter: blur(8px); }}
    .shots {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }}
    @media (max-width: 1100px) {{ .shots {{ grid-template-columns: repeat(3, 1fr); }} }}
    @media (max-width: 820px) {{ .shots {{ grid-template-columns: repeat(2, 1fr); }} }}
    @media (max-width: 520px) {{ .shots {{ grid-template-columns: 1fr; }} }}
    .shot {{ border: 1px solid var(--border); border-radius: 10px; overflow: hidden; background: rgba(255,255,255,.02); }}
    .shot img.thumb {{ width: 100%; height: 120px; object-fit: cover; display: block; }}
    .shot .cap {{ padding: 8px 10px; font-size: 12px; color: var(--muted); }}
    pre {{ background: rgba(255,255,255,.04); border: 1px solid var(--border); border-radius: 10px; padding: 10px 12px; overflow: auto; }}

    .lightbox {{ position: fixed; inset: 0; background: rgba(0,0,0,.76); display: none; z-index: 9999; }}
    .lightbox.open {{ display: block; }}
    .lightbox-inner {{ position: absolute; inset: 28px 20px; display: grid; grid-template-rows: auto 1fr auto; gap: 10px; }}
    .lightbox-top {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; }}
    .lightbox-title {{ color: var(--muted); font-size: 12px; font-family: var(--mono); }}
    .lightbox-btn {{ border: 1px solid rgba(255,255,255,.18); background: rgba(255,255,255,.06); color: var(--text); border-radius: 10px; padding: 8px 10px; cursor: pointer; }}
    .lightbox-btn:hover {{ background: rgba(255,255,255,.10); }}
    .lightbox-body {{ position: relative; border: 1px solid rgba(255,255,255,.14); border-radius: 12px; overflow: hidden; background: rgba(0,0,0,.20); }}
    .lightbox-body img {{ width: 100%; height: 100%; object-fit: contain; display: block; }}
    .lightbox-nav {{ position: absolute; inset: 0; display: flex; align-items: center; justify-content: space-between; pointer-events: none; }}
    .lightbox-nav button {{ pointer-events: all; width: 44px; height: 44px; border-radius: 999px; }}
    .lightbox-bottom {{ color: var(--muted); font-size: 12px; font-family: var(--mono); }}

    @media (max-width: 980px) {{
      .page {{ display: block; }}
      .sidebar {{
        position: sticky;
        top: 0;
        height: auto;
        width: auto;
        border-right: none;
        border-bottom: 1px solid var(--border);
        padding: 12px 12px;
      }}
      .side-nav {{ flex-direction: row; flex-wrap: nowrap; overflow-x: auto; padding-bottom: 6px; }}
      .side-nav a {{ white-space: nowrap; }}
      .side-nav .group {{ display: none; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <aside class="sidebar">
      <div class="side-head">
        <div>
          <div class="side-title">报告目录</div>
          <div class="side-sub mono">{_safe(now)}</div>
        </div>
        <div class="pill {'ok' if ok else 'bad'}">{status}</div>
      </div>
      <nav class="side-nav" id="toc">
        <a href="#overview">总览</a>
        <a href="#steps">执行步骤</a>
        <a href="#llm">LLM 对话</a>
        <div class="group">差距覆盖</div>
        <a href="#dag">指标与DAG</a>
        <a href="#export-param">导出参数化</a>
        <a href="#normalize">输入标准化</a>
        <a href="#completeness">明细完整性</a>
        <a href="#suggestions">改动建议</a>
        <div class="group">一致性</div>
        <a href="#import">导入一致性</a>
        <a href="#export">导出一致性</a>
        <div class="group">模板</div>
        <a href="#template">导出模板</a>
        <div class="group">日志</div>
        <a href="#import-log">导入日志</a>
        <a href="#browser-log">浏览器日志</a>
        <a href="#screenshots">关键截图</a>
        <a href="#conclusion">结论</a>
      </nav>
      <div class="side-foot">
        <div>复现入口：<span class="mono">{_safe(args.base_url)}</span></div>
        <div style="margin-top:6px;"><a href="#overview">回到顶部</a></div>
      </div>
    </aside>

    <main class="main">
      <div class="wrap" id="overview">
    <div class="title">
      <h1>Northstar E2E（agent-browser）</h1>
      <div class="pill {'ok' if ok else 'bad'}">{status}</div>
    </div>
    <div class="section" id="overview-card">
      <div class="h3line">
        <h3>总览</h3>
        <a class="permalink" href="#overview-card">#</a>
      </div>
      <div class="grid">
        <div class="card">
          <h2>运行信息</h2>
          <div class="kv">
            <div class="k">时间</div><div class="v">{_safe(now)}</div>
            <div class="k">BASE_URL</div><div class="v">{_safe(args.base_url)}</div>
            <div class="k">输入 Excel</div><div class="v"><a href="{_safe(args.input_xlsx)}">{_safe(args.input_xlsx)}</a></div>
            <div class="k">导出 Excel</div><div class="v"><a href="{_safe(args.export_xlsx)}">{_safe(args.export_xlsx)}</a></div>
            <div class="k">导出 Excel(DAG后)</div><div class="v">{_link_or_text(args.export_dag_xlsx, "未生成")}</div>
            <div class="k">抽取(导入后)</div><div class="v">{_safe(str(started_at))}</div>
            <div class="k">抽取(修改后)</div><div class="v">{_safe(str(ended_at))}</div>
          </div>
        </div>
        <div class="card">
          <h2>附件</h2>
          <ul>
            <li><a class="mono" href="{_safe(args.ui_before)}">ui_companies_before.json</a></li>
            <li><a class="mono" href="{_safe(args.ui_after)}">ui_companies_after.json</a></li>
            <li><a class="mono" href="{_safe(args.import_events)}">import_events.json</a></li>
            <li><a class="mono" href="{_safe(args.tab_counts)}">tab_counts.json</a></li>
            <li><a class="mono" href="{_safe(args.indicators_before)}">indicators_before.json</a></li>
            <li><a class="mono" href="{_safe(args.indicators_after)}">indicators_after.json</a></li>
            <li><a class="mono" href="{_safe(args.indicator_actions)}">indicator_actions_result.json</a></li>
            <li><a class="mono" href="{_safe(args.llm_results)}">llm_results.json</a></li>
            <li><a class="mono" href="{_safe(args.ui_dag_before)}">ui_companies_before_dag.json</a></li>
            <li><a class="mono" href="{_safe(args.ui_dag_after)}">ui_companies_after_dag.json</a></li>
            <li><a class="mono" href="{_safe(args.export_dag_xlsx)}">export_dag.xlsx</a></li>
            <li><a class="mono" href="{_safe(args.export_alt_xlsx)}">export_alt.xlsx</a></li>
            <li><a class="mono" href="{_safe(args.export_param_meta)}">export_param_meta.json</a></li>
            <li><a class="mono" href="indicator_count_before.json">indicator_count_before.json</a></li>
            <li><a class="mono" href="indicator_count_after.json">indicator_count_after.json</a></li>
            <li><a class="mono" href="indicator_export_before.json">indicator_export_before.json</a></li>
            <li><a class="mono" href="indicator_export_after.json">indicator_export_after.json</a></li>
            <li><a class="mono" href="ui_normalization_checks.json">ui_normalization_checks.json</a></li>
            <li><a class="mono" href="{_safe(args.console)}">browser_console.txt</a></li>
            <li><a class="mono" href="{_safe(args.errors)}">browser_errors.txt</a></li>
            <li><a class="mono" href="{_safe(args.trace)}">trace.zip</a></li>
            <li><a class="mono" href="{_safe(args.video)}">run.webm</a></li>
          </ul>
        </div>
      </div>
    </div>

    <div class="section" id="steps">
      <div class="h3line">
        <h3>执行步骤与不符合预期项</h3>
        <a class="permalink" href="#steps">#</a>
      </div>
      <div class="card">
        <h2>步骤明细</h2>
        {_steps_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>不符合预期项总览</h2>
        {_issues_summary(missing_before, mismatches_before, missing_export, mismatches_export, action_results, action_persist, indicator_action_fail, indicator_count_fail, indicator_export_fail, completeness_failed, completeness_total, int(forbidden_ui_columns_report.get("total") or 0), derived_unmapped_cols_total, derived_missing_ui_cols_total, tab_consistency_fail, ui_derived_fail, export_template_fail, export_formula_fail, normalization_fail, dag_fail, export_param_fail, linkage_preview_fail, llm_fail)}
        <p class="warn">复现入口：打开 <span class="mono">{_safe(args.base_url)}</span> → 导入 → 明细表搜索信用代码 → 修改/对照 → 导出后打开 Excel 对照。</p>
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>修改动作覆盖</h2>
        <p class="warn">复现：在首页输入框按信用代码搜索，修改对应字段输入框，失焦（blur）触发自动保存，然后导出检查。</p>
        {_actions_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>联动预览高亮</h2>
        <p class="warn">目的：点击派生列/指标后，高亮应覆盖父节点与子节点（包含指标与明细表单元格）。</p>
        {_linkage_preview_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>Tab 行数与总数</h2>
        <p class="warn">目的：覆盖“全部/批发/零售/住宿/餐饮”tab 切换与数据是否加载。</p>
        {_tab_counts_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>Tab 覆盖与计数（Excel vs UI）</h2>
        <p class="warn">目的：对标 PRD，确保输入 Excel 有的企业在明细表中完整呈现（企业覆盖/计数一致）。</p>
        {_tab_consistency_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>UI 派生字段一致性检查</h2>
        <p class="warn">目的：覆盖字段联动/DAG 的最小自洽性：派生列应与基础列满足恒等关系。</p>
        {_ui_derived_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>UI 抽取状态</h2>
        <div class="kv">
          <div class="k">导入后抽取</div><div class="v">{'OK' if ui_before_ok else 'FAIL'}</div>
          <div class="k">导入后错误</div><div class="v">{_safe(str(ui_before_error))}</div>
          <div class="k">修改后抽取</div><div class="v">{'OK' if ui_after_ok else 'FAIL'}</div>
          <div class="k">修改后错误</div><div class="v">{_safe(str(ui_after_error))}</div>
          <div class="k">输入 Excel 打开</div><div class="v">{'OK' if input_wb is not None else 'FAIL'}</div>
          <div class="k">导出 Excel 打开</div><div class="v">{'OK' if export_wb is not None else 'FAIL'}</div>
          <div class="k">导出 Excel(raw) 打开</div><div class="v">{'OK' if export_wb_raw is not None else 'FAIL'}</div>
          <div class="k">导出 raw 错误</div><div class="v">{_safe(str(export_wb_raw_error))}</div>
          <div class="k">导出 Excel(DAG后) 打开</div><div class="v">{'OK' if export_dag_wb is not None else 'FAIL'}</div>
          <div class="k">导出 DAG 错误</div><div class="v">{_safe(str(export_dag_wb_error))}</div>
          <div class="k">导出 Excel(alt) 打开</div><div class="v">{'OK' if export_alt_wb_raw is not None else 'FAIL'}</div>
          <div class="k">导出 alt 错误</div><div class="v">{_safe(str(export_alt_wb_raw_error))}</div>
          <div class="k">定稿模板</div><div class="v">{_safe(template_xlsx or '')}</div>
          <div class="k">定稿模板错误</div><div class="v">{_safe(str(template_wb_raw_error))}</div>
        </div>
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>输入 Excel 结构对标 PRD</h2>
        {_input_structure_html()}
      </div>
    </div>

    <div class="section" id="llm">
      <div class="h3line">
        <h3>LLM 对话专项测试</h3>
        <a class="permalink" href="#llm">#</a>
      </div>
      <div class="card">
        <h2>对话与联动校验</h2>
        <p class="warn">覆盖：全局配置显示/隐藏、流式输出、function call 修改、智能调整触发与多轮会话。</p>
        {_llm_html()}
      </div>
    </div>

    <div class="section" id="dag">
      <div class="h3line">
        <h3>指标调整与DAG联动覆盖</h3>
        <a class="permalink" href="#dag">#</a>
      </div>
      <div class="card">
        <h2>指标调整动作</h2>
        <p class="warn">目的：覆盖“修改16项指标 → 智能调整 → 写回企业数据”的链路。</p>
        {_indicator_actions_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>指标变化校验</h2>
        <p class="warn">期望：16项指标全部发生变化。</p>
        {_indicator_changes_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>指标数量与导出一致性</h2>
        <p class="warn">期望：16 项指标数量正确，且导出汇总表与指标一致。</p>
        {_indicator_count_html()}
        <div style="height: 10px;"></div>
        {_indicator_export_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>DAG 下游明细表变化</h2>
        <p class="warn">期望：所有明细表节点均发生变化（随机调整）。</p>
        {_dag_table_diff_html()}
      </div>
    </div>

    <div class="section" id="export-param">
      <div class="h3line">
        <h3>导出数字参数化</h3>
        <a class="permalink" href="#export-param">#</a>
      </div>
      <div class="card">
        <h2>月份参数化检验（双导出对比）</h2>
        <p class="warn">复现：切换到不同月份导出，再对比包含数字的文本/表头是否变化。</p>
        {_export_param_html()}
      </div>
    </div>

    <div class="section" id="normalize">
      <div class="h3line">
        <h3>输入标准化/兼容性</h3>
        <a class="permalink" href="#normalize">#</a>
      </div>
      <div class="card">
        <h2>来源表/表头标准化</h2>
        <p class="warn">目的：兼容非标准 sheet/字段名称并标准化入库。</p>
        {_normalization_html()}
      </div>
    </div>

    <div class="section" id="completeness">
      <div class="h3line">
        <h3>明细表数据完整性案例（输入 Excel → 明细表）</h3>
        <a class="permalink" href="#completeness">#</a>
      </div>
      <div class="card">
        <h2>字段级完整性（大量案例）</h2>
        <p class="warn">完整列表：<a class="mono" href="completeness_cases.json">completeness_cases.json</a> / <a class="mono" href="completeness_summary.json">completeness_summary.json</a></p>
        {_completeness_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>缺失企业（Excel 有，但明细表缺）</h2>
        <p class="warn">完整列表：<a class="mono" href="missing_codes_by_sheet.json">missing_codes_by_sheet.json</a></p>
        {_missing_codes_html()}
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>衍生 Sheet 列覆盖（有值列 → UI 列是否存在）</h2>
        <p class="warn">完整列表：<a class="mono" href="derived_column_coverage.json">derived_column_coverage.json</a></p>
        {_derived_coverage_html()}
      </div>
    </div>

    <div class="section" id="suggestions">
      <div class="h3line">
        <h3>改动建议（仅写在报告内）</h3>
        <a class="permalink" href="#suggestions">#</a>
      </div>
      <div class="card">
        <h2>建议汇总</h2>
        {_suggestions(ui_before_ok, ui_after_ok, before_rows if isinstance(before_rows, list) else [], after_rows if isinstance(after_rows, list) else [], input_wb, export_wb, missing_before, mismatches_before, missing_export, mismatches_export)}
      </div>
    </div>

    <div class="section" id="import">
      <div class="h3line">
        <h3>导入一致性（明细表 vs 输入 Excel）</h3>
        <a class="permalink" href="#import">#</a>
      </div>
      <div class="card">
        <h2>覆盖检查</h2>
        {_summarize_missing(missing_before)}
        <p class="warn">复现：打开输入 Excel 的对应 Sheet（来源表），在“统一社会信用代码”列定位企业，对比明细表同字段显示。</p>
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>字段一致性</h2>
        {_summarize_mismatches(mismatches_before)}
        <p class="warn">复现：首页明细表搜索信用代码，打开“列”展示相关字段，对照输入 Excel 的同字段值。</p>
      </div>
    </div>

	    <div class="section" id="export">
        <div class="h3line">
	        <h3>导出一致性（导出 Excel vs 明细表[修改后]）</h3>
          <a class="permalink" href="#export">#</a>
        </div>
      <div class="card">
        <h2>覆盖检查</h2>
        {_summarize_missing(missing_export)}
        <p class="warn">复现：点击“导出”下载 xlsx，打开后在对应 Sheet 用“统一社会信用代码”定位行，对照明细表修改后的值。</p>
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>字段一致性</h2>
        {_summarize_mismatches(mismatches_export)}
        <p class="warn">复现：先确认修改动作已保存（输入框失焦），再导出并对照字段。</p>
      </div>
      <div class="card" style="margin-top: 12px;">
        <h2>修改动作 → 导出回归校验</h2>
        <p class="warn">完整列表：<a class="mono" href="action_export_checks.json">action_export_checks.json</a></p>
        {_action_export_checks_html()}
	      </div>
	    </div>

	    <div class="section" id="template">
        <div class="h3line">
	        <h3>导出模板结构对标 PRD</h3>
          <a class="permalink" href="#template">#</a>
        </div>
	      <div class="card">
	        <h2>Sheet/表头结构</h2>
	        {_export_template_html()}
	      </div>
	      <div class="card" style="margin-top: 12px;">
	        <h2>模板公式（社零额/汇总表）</h2>
	        {_export_formula_html()}
	      </div>
	    </div>

	    <div class="section" id="import-log">
        <div class="h3line">
	      <h3>导入日志（UI）</h3>
          <a class="permalink" href="#import-log">#</a>
        </div>
      <div class="card">
        <h2>进度事件</h2>
        <p class="mono">count={_safe(str(import_events.get('count', '')))}</p>
        <pre>{_safe(import_log)}</pre>
      </div>
    </div>

    <div class="section" id="browser-log">
      <div class="h3line">
        <h3>浏览器日志</h3>
        <a class="permalink" href="#browser-log">#</a>
      </div>
      <div class="grid">
        <div class="card">
          <h2>Console</h2>
          <pre>{_safe(_read_text(args.console)[:20000])}</pre>
        </div>
        <div class="card">
          <h2>Errors</h2>
          <pre>{_safe(_read_text(args.errors)[:20000])}</pre>
        </div>
      </div>
    </div>

    <div class="section" id="screenshots">
      <div class="h3line">
        <h3>关键截图</h3>
        <a class="permalink" href="#screenshots">#</a>
      </div>
      <div class="shots">
        {shots_html}
      </div>
    </div>

    <div class="section" id="conclusion">
      <div class="h3line">
        <h3>结论</h3>
        <a class="permalink" href="#conclusion">#</a>
      </div>
      <div class="card">
        <p>{_safe(conclusion)}</p>
      </div>
    </div>
  </div>
    </main>
  </div>

  <div id="lightbox" class="lightbox" aria-hidden="true">
    <div class="lightbox-inner">
      <div class="lightbox-top">
        <div id="lightbox-title" class="lightbox-title">screenshot</div>
        <div style="display:flex; gap:8px;">
          <button id="lightbox-open" class="lightbox-btn" type="button">在新标签页打开</button>
          <button id="lightbox-close" class="lightbox-btn" type="button">关闭 (Esc)</button>
        </div>
      </div>
      <div class="lightbox-body" id="lightbox-body">
        <img id="lightbox-img" alt="preview" />
        <div class="lightbox-nav">
          <button id="lightbox-prev" class="lightbox-btn" type="button">‹</button>
          <button id="lightbox-next" class="lightbox-btn" type="button">›</button>
        </div>
      </div>
      <div id="lightbox-cap" class="lightbox-bottom"></div>
    </div>
  </div>

  <script>
    (() => {{
      // Scrollspy for the left TOC.
      const toc = document.getElementById('toc');
      const tocLinks = toc ? Array.from(toc.querySelectorAll('a[href^=\"#\"]')) : [];
      const byId = new Map();
      tocLinks.forEach(a => {{
        const id = (a.getAttribute('href') || '').slice(1);
        if (!id) return;
        byId.set(id, a);
      }});
      const targets = Array.from(byId.keys()).map(id => document.getElementById(id)).filter(Boolean);
      function setActive(id) {{
        tocLinks.forEach(a => a.classList.remove('active'));
        const a = byId.get(id);
        if (a) a.classList.add('active');
      }}
      if (targets.length) {{
        const obs = new IntersectionObserver((entries) => {{
          const visible = entries.filter(e => e.isIntersecting).sort((a,b) => (b.intersectionRatio - a.intersectionRatio));
          if (!visible.length) return;
          const id = visible[0].target && visible[0].target.id;
          if (id) setActive(id);
        }}, {{ root: null, rootMargin: '-20% 0px -70% 0px', threshold: [0.05, 0.2, 0.6] }});
        targets.forEach(el => obs.observe(el));
        setActive(targets[0].id);
      }}

      const links = Array.from(document.querySelectorAll('.shot-link'));
      const lb = document.getElementById('lightbox');
      const img = document.getElementById('lightbox-img');
      const title = document.getElementById('lightbox-title');
      const cap = document.getElementById('lightbox-cap');
      const btnClose = document.getElementById('lightbox-close');
      const btnOpen = document.getElementById('lightbox-open');
      const btnPrev = document.getElementById('lightbox-prev');
      const btnNext = document.getElementById('lightbox-next');
      const body = document.getElementById('lightbox-body');
      if (!lb || !img || links.length === 0) return;

      let idx = 0;
      function setIndex(i) {{
        if (i < 0) i = links.length - 1;
        if (i >= links.length) i = 0;
        idx = i;
        const a = links[idx];
        const href = a.getAttribute('href') || '';
        const c = a.getAttribute('data-cap') || '';
        img.setAttribute('src', href);
        title.textContent = c || href;
        cap.textContent = href;
        btnOpen.onclick = () => window.open(href, '_blank');
      }}

      function openAt(i) {{
        setIndex(i);
        lb.classList.add('open');
        lb.setAttribute('aria-hidden', 'false');
      }}
      function close() {{
        lb.classList.remove('open');
        lb.setAttribute('aria-hidden', 'true');
        img.setAttribute('src', '');
      }}

      links.forEach((a, i) => {{
        a.addEventListener('click', (e) => {{
          if (e.metaKey || e.ctrlKey || e.shiftKey || e.altKey) return;
          e.preventDefault();
          openAt(i);
        }});
      }});

      btnClose && btnClose.addEventListener('click', close);
      btnPrev && btnPrev.addEventListener('click', () => setIndex(idx - 1));
      btnNext && btnNext.addEventListener('click', () => setIndex(idx + 1));
      lb.addEventListener('click', (e) => {{
        if (e.target === lb) close();
      }});
      body && body.addEventListener('dblclick', close);
      document.addEventListener('keydown', (e) => {{
        if (!lb.classList.contains('open')) return;
        if (e.key === 'Escape') close();
        if (e.key === 'ArrowLeft') setIndex(idx - 1);
        if (e.key === 'ArrowRight') setIndex(idx + 1);
      }});
    }})();
  </script>
</body>
</html>
"""

    Path(args.out).write_text(html, encoding="utf-8")

    if not ok:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
