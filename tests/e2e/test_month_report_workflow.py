"""
月报端到端一致性测试

目标：完全模拟
- 12月月报（预估）.xlsx 导入
- 仪表盘微调（修改企业数据）
- 按 12月月报（定）.xlsx 模板导出
- 导出的 Excel 与期望文件一致（包含微调后的数据）

说明：
- 该测试依赖本地真实文件（不内置在仓库）。
- 通过环境变量指定文件路径：
  - NS_MONTH_REPORT_ESTIMATE_XLSX：12月月报（预估）.xlsx
  - NS_MONTH_REPORT_TEMPLATE_XLSX：12月月报（定）.xlsx
  - NS_MONTH_REPORT_EXPECTED_XLSX：期望导出结果（已包含同样微调）
"""

import os
import tempfile
import requests
import pytest
import numbers
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from pathlib import Path

BASE_URL = "http://localhost:18080"


def _env_path(key: str) -> str:
    v = os.environ.get(key, "").strip()
    if not v:
        return ""
    return v


def _assert_file(path: str, key: str):
    if not path:
        pytest.skip(f"跳过：缺少文件路径 {key}")
    if not os.path.exists(path):
        pytest.skip(f"跳过：文件不存在 {key}={path}")


def _upload(session: requests.Session, xlsx_path: str) -> str:
    with open(xlsx_path, "rb") as f:
        files = {
            "file": (
                os.path.basename(xlsx_path),
                f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        res = session.post(f"{BASE_URL}/api/v1/import/upload", files=files)
    assert res.status_code == 200, res.text
    payload = res.json()
    assert payload["code"] == 0, payload
    return payload["data"]["fileId"]


def _execute_report_import(session: requests.Session, file_id: str, month: int):
    res = session.post(
        f"{BASE_URL}/api/v1/import/{file_id}/execute-report",
        json={"month": month, "overrides": {}},
    )
    assert res.status_code == 200, res.text
    payload = res.json()
    assert payload["code"] == 0, payload
    assert payload["data"]["importedCount"] > 0, payload


def _export(session: requests.Session) -> str:
    res = session.post(
        f"{BASE_URL}/api/v1/export",
        json={"format": "xlsx", "includeIndicators": True, "includeChanges": True},
    )
    assert res.status_code == 200, res.text
    payload = res.json()
    assert payload["code"] == 0, payload
    download_url = payload["data"]["downloadUrl"]

    dl = session.get(f"{BASE_URL}{download_url}")
    assert dl.status_code == 200, dl.text
    fd, out = tempfile.mkstemp(suffix=".xlsx")
    os.write(fd, dl.content)
    os.close(fd)
    return out


def _normalize_value(v):
    if v is None:
        return None
    if isinstance(v, numbers.Number):
        return round(float(v), 10)
    return v


def _compare_workbooks(expected_path: str, actual_path: str):
    ew = load_workbook(expected_path, data_only=False)
    aw = load_workbook(actual_path, data_only=False)

    assert aw.sheetnames == ew.sheetnames, f"sheetnames 不一致: {aw.sheetnames} != {ew.sheetnames}"

    for name in ew.sheetnames:
        es = ew[name]
        ac = aw[name]

        assert ac.max_row == es.max_row, f"{name} max_row 不一致: {ac.max_row} != {es.max_row}"
        assert ac.max_column == es.max_column, f"{name} max_column 不一致: {ac.max_column} != {es.max_column}"

        e_merges = sorted([str(r) for r in es.merged_cells.ranges])
        a_merges = sorted([str(r) for r in ac.merged_cells.ranges])
        assert a_merges == e_merges, f"{name} merged ranges 不一致"

        for r in range(1, es.max_row + 1):
            for c in range(1, es.max_column + 1):
                ev = _normalize_value(es.cell(row=r, column=c).value)
                av = _normalize_value(ac.cell(row=r, column=c).value)
                if ev != av:
                    raise AssertionError(f"{name}!R{r}C{c} value 不一致: {av!r} != {ev!r}")


def _default_prd_xlsx(name: str) -> str:
    p = Path(__file__).resolve()
    for d in [p.parent] + list(p.parents):
        if (d / "go.mod").exists():
            return str(d / "prd" / name)
    raise RuntimeError(f"repo root not found from: {p}")

def _rate_percent(cur: float, last: float) -> float:
    if last == 0:
        return -100.0
    return _round_half_up((cur / last - 1) * 100.0, 2)


def _round_half_up(v: float, digits: int) -> float:
    q = Decimal("1").scaleb(-digits)  # 10^-digits
    return float(Decimal(str(v)).quantize(q, rounding=ROUND_HALF_UP))


def _num(v) -> float:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return 0.0
    if isinstance(v, numbers.Number):
        return float(v)
    return float(str(v).strip())


def _find_max_data_row(ws, code_col: str, start_row: int = 2, max_scan: int = 50000) -> int:
    r = start_row
    while r <= max_scan:
        v = ws[f"{code_col}{r}"].value
        if v is None or str(v).strip() == "":
            break
        r += 1
    if r == start_row:
        raise AssertionError(f"{ws.title} 没有数据行")
    return r - 1


def _close(a: float, b: float, eps: float = 1e-6) -> bool:
    return abs(a - b) <= eps


def _find_wholesale_retail_row(ws, industry_code: str, sales_last: float, sales_last_cum: float, retail_last: float, retail_last_cum: float) -> int:
    max_row = _find_max_data_row(ws, "C", 2)
    for r in range(2, max_row + 1):
        code = str(ws[f"C{r}"].value or "").strip()
        if code != str(industry_code).strip():
            continue
        if not _close(_num(ws[f"E{r}"].value), sales_last):
            continue
        if not _close(_num(ws[f"H{r}"].value), sales_last_cum):
            continue
        if not _close(_num(ws[f"K{r}"].value), retail_last):
            continue
        if not _close(_num(ws[f"N{r}"].value), retail_last_cum):
            continue
        return r
    raise AssertionError(f"{ws.title} 找不到匹配行: industryCode={industry_code}")


def _sum_wholesale_retail(ws):
    max_row = _find_max_data_row(ws, "C", 2)
    sums = {
        "sales_cur": 0.0,
        "sales_last": 0.0,
        "sales_cur_cum": 0.0,
        "sales_last_cum": 0.0,
        "retail_cur": 0.0,
        "retail_last": 0.0,
        "retail_cur_cum": 0.0,
        "retail_last_cum": 0.0,
        "max_row": max_row,
    }
    for r in range(2, max_row + 1):
        sums["sales_cur"] += _num(ws[f"D{r}"].value)
        sums["sales_last"] += _num(ws[f"E{r}"].value)
        sums["sales_cur_cum"] += _num(ws[f"G{r}"].value)
        sums["sales_last_cum"] += _num(ws[f"H{r}"].value)
        sums["retail_cur"] += _num(ws[f"J{r}"].value)
        sums["retail_last"] += _num(ws[f"K{r}"].value)
        sums["retail_cur_cum"] += _num(ws[f"M{r}"].value)
        sums["retail_last_cum"] += _num(ws[f"N{r}"].value)
    return sums


def _sum_accommodation_catering(ws):
    max_row = _find_max_data_row(ws, "C", 2)
    sums = {
        "sales_cur": 0.0,
        "sales_last": 0.0,
        "sales_cur_cum": 0.0,
        "sales_last_cum": 0.0,
        "retail_cur": 0.0,
        "retail_last": 0.0,
        "retail_cur_cum": 0.0,
        "retail_last_cum": 0.0,
        "max_row": max_row,
    }
    for r in range(2, max_row + 1):
        sums["sales_cur"] += _num(ws[f"D{r}"].value)
        sums["sales_last"] += _num(ws[f"E{r}"].value)
        sums["sales_cur_cum"] += _num(ws[f"G{r}"].value)
        sums["sales_last_cum"] += _num(ws[f"H{r}"].value)
        sums["retail_cur"] += _num(ws[f"V{r}"].value)
        sums["retail_last"] += _num(ws[f"W{r}"].value)
        sums["retail_cur_cum"] += _num(ws[f"X{r}"].value)
        sums["retail_last_cum"] += _num(ws[f"Y{r}"].value)
    return sums


def _rewrite_totals_wholesale_retail(ws, sums):
    sum_row = sums["max_row"] + 1
    growth_row = sums["max_row"] + 2
    ws[f"D{sum_row}"].value = sums["sales_cur"]
    ws[f"E{sum_row}"].value = sums["sales_last"]
    ws[f"G{sum_row}"].value = sums["sales_cur_cum"]
    ws[f"H{sum_row}"].value = sums["sales_last_cum"]
    ws[f"J{sum_row}"].value = sums["retail_cur"]
    ws[f"K{sum_row}"].value = sums["retail_last"]
    ws[f"M{sum_row}"].value = sums["retail_cur_cum"]
    ws[f"N{sum_row}"].value = sums["retail_last_cum"]
    ws[f"E{growth_row}"].value = _rate_percent(sums["sales_cur"], sums["sales_last"])
    ws[f"H{growth_row}"].value = _rate_percent(sums["sales_cur_cum"], sums["sales_last_cum"])


def _rewrite_totals_accommodation_catering(ws, sums):
    sum_row = sums["max_row"] + 1
    growth_row = sums["max_row"] + 2
    ws[f"D{sum_row}"].value = sums["sales_cur"]
    ws[f"E{sum_row}"].value = sums["sales_last"]
    ws[f"G{sum_row}"].value = sums["sales_cur_cum"]
    ws[f"H{sum_row}"].value = sums["sales_last_cum"]
    ws[f"V{sum_row}"].value = sums["retail_cur"]
    ws[f"W{sum_row}"].value = sums["retail_last"]
    ws[f"X{sum_row}"].value = sums["retail_cur_cum"]
    ws[f"Y{sum_row}"].value = sums["retail_last_cum"]
    ws[f"E{growth_row}"].value = _rate_percent(sums["sales_cur"], sums["sales_last"])
    ws[f"H{growth_row}"].value = _rate_percent(sums["sales_cur_cum"], sums["sales_last_cum"])


def _rewrite_overall_retail_area_on_wholesale(wb, wh, re, acc, cat):
    ws = wb["批发"]
    wh_max = wh["max_row"]
    sum_row = wh_max + 1
    growth_row = wh_max + 2
    acc_row = growth_row + 1
    cat_row = growth_row + 2
    total_row = growth_row + 3
    total_growth_row = growth_row + 4

    ws[f"J{growth_row}"].value = re["retail_cur"]
    ws[f"K{growth_row}"].value = re["retail_last"]
    ws[f"M{growth_row}"].value = re["retail_cur_cum"]
    ws[f"N{growth_row}"].value = re["retail_last_cum"]

    ws[f"J{acc_row}"].value = acc["retail_cur"]
    ws[f"K{acc_row}"].value = acc["retail_last"]
    ws[f"M{acc_row}"].value = acc["retail_cur_cum"]
    ws[f"N{acc_row}"].value = acc["retail_last_cum"]

    ws[f"J{cat_row}"].value = cat["retail_cur"]
    ws[f"K{cat_row}"].value = cat["retail_last"]
    ws[f"M{cat_row}"].value = cat["retail_cur_cum"]
    ws[f"N{cat_row}"].value = cat["retail_last_cum"]

    overall_cur = wh["retail_cur"] + re["retail_cur"] + acc["retail_cur"] + cat["retail_cur"]
    overall_last = wh["retail_last"] + re["retail_last"] + acc["retail_last"] + cat["retail_last"]
    overall_cur_cum = wh["retail_cur_cum"] + re["retail_cur_cum"] + acc["retail_cur_cum"] + cat["retail_cur_cum"]
    overall_last_cum = wh["retail_last_cum"] + re["retail_last_cum"] + acc["retail_last_cum"] + cat["retail_last_cum"]

    ws[f"J{total_row}"].value = overall_cur
    ws[f"K{total_row}"].value = overall_last
    ws[f"M{total_row}"].value = overall_cur_cum
    ws[f"N{total_row}"].value = overall_last_cum

    ws[f"K{total_growth_row}"].value = _rate_percent(overall_cur, overall_last)
    ws[f"N{total_growth_row}"].value = _rate_percent(overall_cur_cum, overall_last_cum)

    ws[f"J{sum_row}"].value = wh["retail_cur"]
    ws[f"K{sum_row}"].value = wh["retail_last"]
    ws[f"M{sum_row}"].value = wh["retail_cur_cum"]
    ws[f"N{sum_row}"].value = wh["retail_last_cum"]


def _rewrite_fixed_summary_sheet(wb, wh, re, acc, cat):
    summary = wb["汇总表（定）"]
    ws = wb["批发"]
    wh_max = wh["max_row"]
    overall_row = wh_max + 5
    overall_cur = _num(ws[f"J{overall_row}"].value)
    overall_last = _num(ws[f"K{overall_row}"].value)
    overall_cur_cum = _num(ws[f"M{overall_row}"].value)
    overall_last_cum = _num(ws[f"N{overall_row}"].value)

    summary["G4"].value = overall_cur / 10
    summary["H4"].value = overall_last / 10
    summary["I4"].value = overall_cur_cum / 10
    summary["J4"].value = overall_last_cum / 10

    summary["K4"].value = _round_half_up(_num(wb["批发"][f"E{wh_max+2}"].value), 1)
    summary["L4"].value = _round_half_up(_num(wb["批发"][f"H{wh_max+2}"].value), 1)
    re_max = re["max_row"]
    summary["M4"].value = _round_half_up(_num(wb["零售"][f"E{re_max+2}"].value), 1)
    summary["N4"].value = _round_half_up(_num(wb["零售"][f"H{re_max+2}"].value), 1)
    acc_max = acc["max_row"]
    summary["O4"].value = _round_half_up(_num(wb["住宿"][f"E{acc_max+2}"].value), 1)
    summary["P4"].value = _round_half_up(_num(wb["住宿"][f"H{acc_max+2}"].value), 1)
    cat_max = cat["max_row"]
    summary["Q4"].value = _round_half_up(_num(wb["餐饮"][f"E{cat_max+2}"].value), 1)
    summary["R4"].value = _round_half_up(_num(wb["餐饮"][f"H{cat_max+2}"].value), 1)

    summary["S4"].value = _round_half_up((overall_cur / overall_last - 1) * 100.0 if overall_last != 0 else -100.0, 1)
    summary["T4"].value = _round_half_up((overall_cur_cum / overall_last_cum - 1) * 100.0 if overall_last_cum != 0 else -100.0, 1)


def _copy_data_rows(wb, a_sheet: str, b_sheet: str, out_sheet: str, code_col: str = "C"):
    a = wb[a_sheet]
    b = wb[b_sheet]
    out = wb[out_sheet]
    a_max = _find_max_data_row(a, code_col, 2)
    b_max = _find_max_data_row(b, code_col, 2)
    out_row = 2
    for r in range(2, a_max + 1):
        for c in range(1, out.max_column + 1):
            out.cell(row=out_row, column=c).value = a.cell(row=r, column=c).value
        out_row += 1
    for r in range(2, b_max + 1):
        for c in range(1, out.max_column + 1):
            out.cell(row=out_row, column=c).value = b.cell(row=r, column=c).value
        out_row += 1


def test_month_report_import_adjust_export_matches_expected(api_client):
    session, _ = api_client
    estimate = _env_path("NS_MONTH_REPORT_ESTIMATE_XLSX") or _default_prd_xlsx("12月月报（预估）.xlsx")
    template = _env_path("NS_MONTH_REPORT_TEMPLATE_XLSX") or _default_prd_xlsx("12月月报（定）.xlsx")
    expected = _env_path("NS_MONTH_REPORT_EXPECTED_XLSX")

    _assert_file(estimate, "NS_MONTH_REPORT_ESTIMATE_XLSX")
    _assert_file(template, "NS_MONTH_REPORT_TEMPLATE_XLSX")

    file_id = _upload(session, estimate)
    _execute_report_import(session, file_id, 12)

    # 微调：选择“非吃穿用/非小微”的批零企业，确保预期更新范围可控
    res = session.get(f"{BASE_URL}/api/v1/companies?page=1&pageSize=200&sortBy=rowNo&sortDir=asc")
    assert res.status_code == 200, res.text
    payload = res.json()
    assert payload["code"] == 0, payload
    items = payload["data"]["items"]
    c = None
    for it in items:
        if it.get("industryType") not in ("wholesale", "retail"):
            continue
        if bool(it.get("isEatWearUse")):
            continue
        if int(it.get("companyScale") or 0) in (3, 4):
            continue
        # 避免模板 key 碰撞：关键字段全为 0 时，无法稳定定位到模板行（后端会跳过这类行写入）。
        if (
            float(it.get("salesLastYearMonth") or 0) == 0
            and float(it.get("salesLastYearCumulative") or 0) == 0
            and float(it.get("retailLastYearMonth") or 0) == 0
            and float(it.get("retailLastYearCumulative") or 0) == 0
        ):
            continue
        c = it
        break
    if c is None:
        pytest.skip("跳过：未找到合适的批零企业用于微调（非吃穿用/非小微）")
    cid = c["id"]
    industry_type = c.get("industryType")
    industry_code = str(c.get("industryCode", "")).strip()
    retail = float(c.get("retailCurrentMonth", 0))
    sales = float(c.get("salesCurrentMonth", 0))
    retail_cum = float(c.get("retailCurrentCumulative", 0))
    sales_cum = float(c.get("salesCurrentCumulative", 0))
    retail_last = float(c.get("retailLastYearMonth", 0))
    sales_last = float(c.get("salesLastYearMonth", 0))
    retail_last_cum = float(c.get("retailLastYearCumulative", 0))
    sales_last_cum = float(c.get("salesLastYearCumulative", 0))

    next_retail = retail + 123.45
    next_sales = max(sales, next_retail + 1)

    res = session.patch(
        f"{BASE_URL}/api/v1/companies/{cid}",
        json={"retailCurrentMonth": next_retail, "salesCurrentMonth": next_sales},
    )
    assert res.status_code == 200, res.text
    payload = res.json()
    assert payload["code"] == 0, payload

    exported = _export(session)

    # 期望文件：优先使用外部提供的“已包含同样微调”的期望文件；否则在模板上按同样规则生成期望。
    if expected:
        _assert_file(expected, "NS_MONTH_REPORT_EXPECTED_XLSX")
        _compare_workbooks(expected, exported)
        return

    # 在模板上生成期望（按后端“月报导出器”同样规则回写汇总区）
    tw = load_workbook(template, data_only=False)
    sheet_name = "批发" if industry_type == "wholesale" else "零售"
    ws = tw[sheet_name]
    row = _find_wholesale_retail_row(ws, industry_code, sales_last, sales_last_cum, retail_last, retail_last_cum)

    # 后端会把当月 delta 同步到累计
    next_sales_cum = sales_cum + (next_sales - sales)
    next_retail_cum = retail_cum + (next_retail - retail)

    ws[f"D{row}"].value = next_sales
    ws[f"G{row}"].value = next_sales_cum
    ws[f"J{row}"].value = next_retail
    ws[f"M{row}"].value = next_retail_cum
    ws[f"F{row}"].value = _rate_percent(next_sales, sales_last)
    ws[f"I{row}"].value = _rate_percent(next_sales_cum, sales_last_cum)
    ws[f"L{row}"].value = _rate_percent(next_retail, retail_last)
    ws[f"O{row}"].value = _rate_percent(next_retail_cum, retail_last_cum)

    wh = _sum_wholesale_retail(tw["批发"])
    re_sum = _sum_wholesale_retail(tw["零售"])
    acc = _sum_accommodation_catering(tw["住宿"])
    cat = _sum_accommodation_catering(tw["餐饮"])
    _rewrite_totals_wholesale_retail(tw["批发"], wh)
    _rewrite_totals_wholesale_retail(tw["零售"], re_sum)
    _rewrite_totals_accommodation_catering(tw["住宿"], acc)
    _rewrite_totals_accommodation_catering(tw["餐饮"], cat)
    _rewrite_overall_retail_area_on_wholesale(tw, wh, re_sum, acc, cat)
    _rewrite_fixed_summary_sheet(tw, wh, re_sum, acc, cat)

    # 总表：按模板行顺序定位并回写（不做“拼接复制”，与后端实现保持一致）
    total = tw["批零总表"]
    total_row = _find_wholesale_retail_row(total, industry_code, sales_last, sales_last_cum, retail_last, retail_last_cum)
    total[f"D{total_row}"].value = next_sales
    total[f"G{total_row}"].value = next_sales_cum
    total[f"J{total_row}"].value = next_retail
    total[f"M{total_row}"].value = next_retail_cum
    total[f"F{total_row}"].value = _rate_percent(next_sales, sales_last)
    total[f"I{total_row}"].value = _rate_percent(next_sales_cum, sales_last_cum)
    total[f"L{total_row}"].value = _rate_percent(next_retail, retail_last)
    total[f"O{total_row}"].value = _rate_percent(next_retail_cum, retail_last_cum)

    fd, expected_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    tw.save(expected_path)

    _compare_workbooks(expected_path, exported)
