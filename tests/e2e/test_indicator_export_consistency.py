import math
from pathlib import Path

import openpyxl

import tests.e2e.run_agent_browser_e2e_report as report


def _build_export_wb(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总表（定）"
    # Fill indicator cells
    ws["G4"].value = 200  # limitAbove_month_value
    ws["S4"].value = 5  # limitAbove_month_rate
    ws["I4"].value = 800  # limitAbove_cumulative_value
    ws["T4"].value = 6  # limitAbove_cumulative_rate
    ws["K4"].value = 7
    ws["L4"].value = 8
    ws["M4"].value = 9
    ws["N4"].value = 10
    ws["O4"].value = 11
    ws["P4"].value = 12
    ws["Q4"].value = 13
    ws["R4"].value = 14
    ws["U4"].value = 15
    ws["V4"].value = 16
    # totalSocial_cumulative_value in 亿元 with 2 decimals
    ws["N10"].value = 12.35
    ws["S10"].value = 17

    path = tmp_path / "export.xlsx"
    wb.save(path)
    return openpyxl.load_workbook(path, data_only=True)


def _indicator_items():
    return [
        {"label": "限上社零额（当月值）", "value": 200, "unit": "万元"},
        {"label": "限上社零额增速（当月）", "value": 5, "unit": "%"},
        {"label": "限上社零额（累计值）", "value": 800, "unit": "万元"},
        {"label": "限上社零额增速（累计）", "value": 6, "unit": "%"},
        {"label": "批发业销售额增速（当月）", "value": 7, "unit": "%"},
        {"label": "批发业销售额增速（累计）", "value": 8, "unit": "%"},
        {"label": "零售业销售额增速（当月）", "value": 9, "unit": "%"},
        {"label": "零售业销售额增速（累计）", "value": 10, "unit": "%"},
        {"label": "住宿业营业额增速（当月）", "value": 11, "unit": "%"},
        {"label": "住宿业营业额增速（累计）", "value": 12, "unit": "%"},
        {"label": "餐饮业营业额增速（当月）", "value": 13, "unit": "%"},
        {"label": "餐饮业营业额增速（累计）", "value": 14, "unit": "%"},
        {"label": "吃穿用增速（当月）", "value": 15, "unit": "%"},
        {"label": "小微企业增速（当月）", "value": 16, "unit": "%"},
        # totalSocial_cumulative_value in 万元 -> export uses 亿元
        {"label": "社零总额（累计值）", "value": 123456, "unit": "万元"},
        {"label": "社零总额增速（累计）", "value": 17, "unit": "%"},
    ]


def test_indicator_export_consistency_ok(tmp_path):
    wb = _build_export_wb(tmp_path)
    items = _indicator_items()
    # adjust export cell for total social value
    expected_yi = report._round_half_up(123456 / 10000.0, 2)
    assert math.isclose(expected_yi, 12.35, abs_tol=1e-6)

    res = report._compare_indicator_values_to_export(items, wb)
    assert res.get("ok") is True
    assert res.get("total") == 16
    assert res.get("mismatch") == 0


def test_indicator_export_consistency_mismatch(tmp_path):
    wb = _build_export_wb(tmp_path)
    items = _indicator_items()
    # introduce mismatch
    items[0]["value"] = 999
    res = report._compare_indicator_values_to_export(items, wb)
    assert res.get("ok") is False
    assert res.get("mismatch") == 1
