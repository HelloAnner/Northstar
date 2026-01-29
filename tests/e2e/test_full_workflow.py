"""
Northstar E2E 完整流程测试
测试完整的数据导入、修改、导出流程

使用真实 Excel 文件（300条企业数据）进行测试
"""
import os
import time
import shutil
import pytest
import requests
from openpyxl import load_workbook

from test_data_generator import generate_company_data, create_excel_file, get_field_mapping


# 测试配置
BASE_URL = "http://localhost:18080"
FIXTURES_DIR = os.path.join(os.path.dirname(__file__), 'fixtures')
TEST_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'e2e-result')


@pytest.fixture(scope="module")
def setup_test_environment():
    """设置测试环境"""
    # 创建目录
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)

    # 生成测试 Excel 文件
    companies = generate_company_data(count=300, seed=42)
    test_file = os.path.join(FIXTURES_DIR, 'test_companies_300.xlsx')
    create_excel_file(companies, test_file)

    yield {
        'test_file': test_file,
        'companies': companies,
        'fixtures_dir': FIXTURES_DIR,
        'output_dir': TEST_OUTPUT_DIR,
    }


@pytest.fixture(scope="module")
def api_session():
    """创建 API 会话"""
    session = requests.Session()

    # 等待服务器启动
    max_retries = 30
    for i in range(max_retries):
        try:
            response = session.get(f"{BASE_URL}/api/v1/indicators")
            if response.status_code == 200:
                break
        except requests.ConnectionError:
            pass
        time.sleep(1)
    else:
        pytest.fail("服务器未能在规定时间内启动")

    yield session

    session.close()


class TestFullWorkflow:
    """完整工作流程测试"""

    file_id = None  # 保存上传后的文件ID
    imported_companies = []  # 保存导入后的企业列表

    def test_01_upload_excel_file(self, api_session, setup_test_environment):
        """测试上传 Excel 文件"""
        test_file = setup_test_environment['test_file']

        # 上传文件
        with open(test_file, 'rb') as f:
            files = {'file': ('test_companies_300.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = api_session.post(f"{BASE_URL}/api/v1/import/upload", files=files)

        assert response.status_code == 200, f"上传失败: {response.text}"

        data = response.json()
        assert data['code'] == 0, f"上传失败: {data.get('message')}"
        assert 'data' in data
        assert 'fileId' in data['data']
        assert 'sheets' in data['data']

        # 保存文件ID
        TestFullWorkflow.file_id = data['data']['fileId']

        print(f"\n上传成功，文件ID: {TestFullWorkflow.file_id}")
        print(f"工作表: {data['data']['sheets']}")

    def test_02_get_columns(self, api_session):
        """测试获取列信息"""
        assert TestFullWorkflow.file_id is not None, "需要先上传文件"

        response = api_session.get(
            f"{BASE_URL}/api/v1/import/{TestFullWorkflow.file_id}/columns",
            params={'sheet': '企业数据'}
        )

        assert response.status_code == 200, f"获取列信息失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        columns = data['data']['columns']
        print(f"\n获取到列信息: {columns}")

        # 验证必要的列存在
        required_columns = ['企业名称', '行业代码', '企业规模', '本期零售额', '上年同期零售额']
        for col in required_columns:
            assert col in columns, f"缺少必要列: {col}"

    def test_03_set_mapping(self, api_session):
        """测试设置字段映射"""
        assert TestFullWorkflow.file_id is not None, "需要先上传文件"

        mapping = get_field_mapping()

        response = api_session.post(
            f"{BASE_URL}/api/v1/import/{TestFullWorkflow.file_id}/mapping",
            json={
                'sheet': '企业数据',
                'mapping': mapping
            }
        )

        assert response.status_code == 200, f"设置映射失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        print("\n字段映射设置成功")

    def test_04_execute_import(self, api_session, setup_test_environment):
        """测试执行导入"""
        assert TestFullWorkflow.file_id is not None, "需要先上传文件"

        response = api_session.post(
            f"{BASE_URL}/api/v1/import/{TestFullWorkflow.file_id}/execute",
            json={
                'sheet': '企业数据',
                'generateHistory': False,
                'currentMonth': 6
            }
        )

        assert response.status_code == 200, f"导入失败: {response.text}"

        data = response.json()
        assert data['code'] == 0, f"导入失败: {data.get('message')}"

        imported_count = data['data']['importedCount']
        expected_count = len(setup_test_environment['companies'])

        print(f"\n导入成功: {imported_count} 条企业数据")
        print(f"预期导入: {expected_count} 条")

        # 允许一定的误差（可能有空行或无效行）
        assert imported_count >= expected_count * 0.95, f"导入数量不足: {imported_count} < {expected_count * 0.95}"

        # 验证指标已计算
        indicators = data['data'].get('indicators')
        if indicators:
            print(f"当月限上社零额增速: {indicators.get('limitAboveMonthRate', 0):.2%}")
            print(f"累计限上社零额增速: {indicators.get('limitAboveCumulativeRate', 0):.2%}")

    def test_05_list_companies(self, api_session):
        """测试获取企业列表"""
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'page': 1, 'pageSize': 50}
        )

        assert response.status_code == 200, f"获取企业列表失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        total = data['data']['total']
        items = data['data']['items']

        print(f"\n企业总数: {total}")
        print(f"当前页数量: {len(items)}")

        # 保存企业列表用于后续测试
        TestFullWorkflow.imported_companies = items

        assert total >= 280, f"导入的企业数量不足: {total}"
        assert len(items) == 50, "分页大小不正确"

    def test_06_modify_single_company(self, api_session):
        """测试修改单个企业数据"""
        assert len(TestFullWorkflow.imported_companies) > 0, "需要先获取企业列表"

        # 选择第一个企业进行修改
        company = TestFullWorkflow.imported_companies[0]
        company_id = company['id']
        original_value = company['retailCurrentMonth']
        new_value = original_value * 1.1  # 增加10%

        print(f"\n修改企业: {company['name']}")
        print(f"原始零售额: {original_value:.2f}")
        print(f"新零售额: {new_value:.2f}")

        response = api_session.patch(
            f"{BASE_URL}/api/v1/companies/{company_id}",
            json={'retailCurrentMonth': new_value}
        )

        assert response.status_code == 200, f"修改失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        # 验证修改成功
        updated_value = data['data']['company']['retailCurrentMonth']
        assert abs(updated_value - new_value) < 0.01, f"修改后的值不正确: {updated_value}"

        print(f"修改成功，新增速: {data['data']['company'].get('monthGrowthRate', 0):.2%}")

    def test_07_batch_modify_companies(self, api_session):
        """测试批量修改企业数据"""
        assert len(TestFullWorkflow.imported_companies) >= 10, "需要足够的企业数据"

        # 选择10个企业进行批量修改
        updates = []
        for company in TestFullWorkflow.imported_companies[1:11]:
            updates.append({
                'id': company['id'],
                'retailCurrentMonth': company['retailCurrentMonth'] * 1.05  # 增加5%
            })

        print(f"\n批量修改 {len(updates)} 个企业")

        response = api_session.patch(
            f"{BASE_URL}/api/v1/companies/batch",
            json={'updates': updates}
        )

        assert response.status_code == 200, f"批量修改失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        updated_count = data['data']['updatedCount']
        assert updated_count == len(updates), f"更新数量不正确: {updated_count}"

        print(f"批量修改成功: {updated_count} 个企业")

    def test_08_verify_indicators_updated(self, api_session):
        """测试验证指标已更新"""
        response = api_session.get(f"{BASE_URL}/api/v1/indicators")

        assert response.status_code == 200, f"获取指标失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        indicators = data['data']

        print("\n当前指标:")
        print(f"  限上社零额(当月): {indicators['limitAboveMonthValue']:.2f} 万元")
        print(f"  限上社零额增速(当月): {indicators['limitAboveMonthRate']:.2%}")
        print(f"  限上社零额(累计): {indicators['limitAboveCumulativeValue']:.2f} 万元")
        print(f"  限上社零额增速(累计): {indicators['limitAboveCumulativeRate']:.2%}")
        print(f"  吃穿用增速(当月): {indicators['eatWearUseMonthRate']:.2%}")
        print(f"  小微企业增速(当月): {indicators['microSmallMonthRate']:.2%}")

        # 验证指标不为0（因为有数据）
        assert indicators['limitAboveMonthValue'] > 0, "当月限上社零额应大于0"
        assert indicators['limitAboveCumulativeValue'] > 0, "累计限上社零额应大于0"

    def test_09_search_companies(self, api_session):
        """测试企业搜索功能"""
        # 搜索包含"有限公司"的企业
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'search': '有限公司', 'page': 1, 'pageSize': 20}
        )

        assert response.status_code == 200, f"搜索失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        items = data['data']['items']
        total = data['data']['total']

        print(f"\n搜索'有限公司'结果: {total} 条")

        # 验证所有结果都包含搜索关键词
        for item in items:
            assert '有限公司' in item['name'], f"搜索结果不正确: {item['name']}"

    def test_10_filter_by_industry(self, api_session):
        """测试按行业筛选"""
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'industry': 'retail', 'page': 1, 'pageSize': 50}
        )

        assert response.status_code == 200, f"筛选失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        items = data['data']['items']
        total = data['data']['total']

        print(f"\n零售业企业: {total} 家")

        # 验证所有结果都是零售业
        for item in items:
            assert item['industryType'] == 'retail', f"行业类型不正确: {item['industryType']}"

    def test_11_smart_optimize_preview(self, api_session):
        """测试智能调整预览"""
        response = api_session.post(
            f"{BASE_URL}/api/v1/optimize/preview",
            json={
                'targetIndicator': 'limitAboveCumulativeRate',
                'targetValue': 0.08,  # 目标8%增速
                'constraints': {
                    'maxIndividualRate': 0.5,
                    'minIndividualRate': 0
                }
            }
        )

        assert response.status_code == 200, f"预览失败: {response.text}"

        data = response.json()
        print(f"\n智能调整预览: {data}")

    def test_12_export_data(self, api_session, setup_test_environment):
        """测试导出数据"""
        response = api_session.post(
            f"{BASE_URL}/api/v1/export",
            json={
                'format': 'xlsx',
                'includeIndicators': True,
                'includeChanges': True
            }
        )

        assert response.status_code == 200, f"导出请求失败: {response.text}"

        data = response.json()
        assert data['code'] == 0, f"导出失败: {data.get('message')}"

        download_url = data['data']['downloadUrl']
        print(f"\n导出下载URL: {download_url}")

        # 下载导出文件
        download_response = api_session.get(f"{BASE_URL}{download_url}")
        assert download_response.status_code == 200, f"下载失败: {download_response.status_code}"

        # 保存导出文件
        output_file = os.path.join(setup_test_environment['output_dir'], 'exported_data.xlsx')
        with open(output_file, 'wb') as f:
            f.write(download_response.content)

        print(f"导出文件已保存: {output_file}")

        # 验证文件大小
        file_size = os.path.getsize(output_file)
        assert file_size > 1000, f"导出文件太小: {file_size} bytes"

        print(f"导出文件大小: {file_size} bytes")

    def test_12b_export_matches_template(self, setup_test_environment):
        """验证导出结构与定稿模板一致（sheet 数量与顺序）"""
        output_file = os.path.join(setup_test_environment['output_dir'], 'exported_data.xlsx')
        assert os.path.exists(output_file), "需要先生成导出文件"

        wb = load_workbook(output_file)
        expected = [
            "批零总表",
            "住餐总表",
            "批发",
            "零售",
            "住宿",
            "餐饮",
            "吃穿用",
            "小微",
            "吃穿用（剔除）",
            "社零额（定）",
            "汇总表（定）",
        ]
        assert wb.sheetnames == expected, f"sheetnames 不一致: {wb.sheetnames}"

    def test_13_reset_companies(self, api_session):
        """测试重置企业数据"""
        # 先获取一些修改过的企业
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'page': 1, 'pageSize': 5}
        )

        data = response.json()
        company_ids = [c['id'] for c in data['data']['items'][:3]]

        print(f"\n重置 {len(company_ids)} 个企业")

        # 执行重置
        response = api_session.post(
            f"{BASE_URL}/api/v1/companies/reset",
            json={'companyIds': company_ids}
        )

        assert response.status_code == 200, f"重置失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        print("重置成功")

    def test_14_reset_all_and_verify(self, api_session):
        """测试重置所有数据并验证"""
        # 重置所有数据
        response = api_session.post(
            f"{BASE_URL}/api/v1/companies/reset",
            json={}
        )

        assert response.status_code == 200, f"重置失败: {response.text}"

        data = response.json()
        assert data['code'] == 0

        print("\n已重置所有企业数据")


class TestDataIntegrity:
    """数据完整性测试"""

    def test_industry_distribution(self, api_session, setup_test_environment):
        """验证行业分布"""
        # 获取所有企业
        all_companies = []
        page = 1
        while True:
            response = api_session.get(
                f"{BASE_URL}/api/v1/companies",
                params={'page': page, 'pageSize': 100}
            )
            data = response.json()
            if data['code'] != 0:
                break

            items = data['data']['items']
            if not items:
                break

            all_companies.extend(items)
            if len(items) < 100:
                break
            page += 1

        if not all_companies:
            pytest.skip("没有企业数据")

        # 统计行业分布
        industry_count = {}
        for company in all_companies:
            industry = company.get('industryType', 'unknown')
            industry_count[industry] = industry_count.get(industry, 0) + 1

        print(f"\n行业分布 (共 {len(all_companies)} 家):")
        for industry, count in sorted(industry_count.items()):
            print(f"  {industry}: {count} ({count/len(all_companies)*100:.1f}%)")

        # 验证四大行业都有数据
        expected_industries = ['wholesale', 'retail', 'accommodation', 'catering']
        for industry in expected_industries:
            assert industry in industry_count, f"缺少 {industry} 行业数据"

    def test_scale_distribution(self, api_session):
        """验证企业规模分布"""
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'page': 1, 'pageSize': 300}
        )

        data = response.json()
        if data['code'] != 0 or not data['data']['items']:
            pytest.skip("没有企业数据")

        companies = data['data']['items']

        # 统计规模分布
        scale_count = {}
        for company in companies:
            scale = company.get('companyScale', 0)
            scale_count[scale] = scale_count.get(scale, 0) + 1

        scale_names = {1: '大型', 2: '中型', 3: '小型', 4: '微型'}
        print(f"\n规模分布 (共 {len(companies)} 家):")
        for scale, count in sorted(scale_count.items()):
            name = scale_names.get(scale, f'规模{scale}')
            print(f"  {name}: {count} ({count/len(companies)*100:.1f}%)")


class TestEdgeCases:
    """边界条件测试"""

    def test_empty_file_upload(self, api_session, setup_test_environment):
        """测试上传空文件"""
        from openpyxl import Workbook

        # 创建空 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "空数据"

        empty_file = os.path.join(setup_test_environment['fixtures_dir'], 'empty.xlsx')
        wb.save(empty_file)

        # 上传
        with open(empty_file, 'rb') as f:
            files = {'file': ('empty.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = api_session.post(f"{BASE_URL}/api/v1/import/upload", files=files)

        assert response.status_code == 200

        print("\n空文件上传测试通过")

    def test_large_value_modification(self, api_session):
        """测试大数值修改"""
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'page': 1, 'pageSize': 1}
        )

        data = response.json()
        if data['code'] != 0 or not data['data']['items']:
            pytest.skip("没有企业数据")

        company = data['data']['items'][0]
        company_id = company['id']

        # 尝试设置一个很大的值
        large_value = 999999999.99

        response = api_session.patch(
            f"{BASE_URL}/api/v1/companies/{company_id}",
            json={'retailCurrentMonth': large_value}
        )

        assert response.status_code == 200
        print(f"\n大数值修改测试通过: {large_value}")

    def test_negative_value_validation(self, api_session):
        """测试负数值校验"""
        response = api_session.get(
            f"{BASE_URL}/api/v1/companies",
            params={'page': 1, 'pageSize': 1}
        )

        data = response.json()
        if data['code'] != 0 or not data['data']['items']:
            pytest.skip("没有企业数据")

        company = data['data']['items'][0]
        company_id = company['id']

        # 尝试设置负数
        response = api_session.patch(
            f"{BASE_URL}/api/v1/companies/{company_id}",
            json={'retailCurrentMonth': -100}
        )

        assert response.status_code == 200

        data = response.json()
        # 应该有校验警告
        validation = data['data']['company'].get('validation', {})
        print(f"\n负数校验结果: {validation}")


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--html=e2e_report.html', '--self-contained-html'])
