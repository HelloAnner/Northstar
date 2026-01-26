"""
测试数据生成器
生成包含几百条企业数据的 Excel 文件用于 E2E 测试
"""
import os
import random
from datetime import datetime
from openpyxl import Workbook
from faker import Faker

# 设置中文 locale
fake = Faker('zh_CN')


# 行业代码定义
INDUSTRY_CODES = {
    'wholesale': ['5111', '5112', '5113', '5121', '5122', '5123', '5131', '5132'],  # 批发业
    'retail': ['5211', '5212', '5213', '5221', '5222', '5223', '5241', '5242', '5243', '5251', '5252'],  # 零售业
    'accommodation': ['6110', '6120', '6130'],  # 住宿业
    'catering': ['6210', '6220', '6230', '6240'],  # 餐饮业
}

# 吃穿用类行业代码
EAT_WEAR_USE_CODES = ['5211', '5212', '5213', '5221', '5222', '5223', '5241', '5242', '5243', '5122', '5123']

# 企业名称后缀
COMPANY_SUFFIXES = [
    '有限公司', '股份有限公司', '集团有限公司', '商贸有限公司',
    '贸易有限公司', '实业有限公司', '科技有限公司', '发展有限公司'
]

# 企业名称前缀（按行业）
COMPANY_PREFIXES = {
    'wholesale': ['华联', '永辉', '中粮', '北大荒', '金龙鱼', '伊利', '蒙牛', '双汇', '三全', '思念'],
    'retail': ['华润万家', '沃尔玛', '家乐福', '大润发', '苏宁', '国美', '京东', '天猫', '唯品会', '拼多多'],
    'accommodation': ['如家', '汉庭', '锦江', '华住', '首旅', '格林豪泰', '亚朵', '全季', '维也纳', '希尔顿'],
    'catering': ['海底捞', '西贝', '呷哺呷哺', '外婆家', '绿茶', '巴奴', '太二', '九毛九', '喜茶', '奈雪'],
}


def generate_credit_code():
    """生成统一社会信用代码（18位）"""
    # 简化版：直接生成18位数字字母混合
    chars = '0123456789ABCDEFGHJKLMNPQRTUWXY'
    return ''.join(random.choices(chars, k=18))


def generate_company_data(count: int = 300, seed: int = 42) -> list:
    """
    生成企业测试数据

    Args:
        count: 生成的企业数量
        seed: 随机种子，确保可重复

    Returns:
        企业数据列表
    """
    random.seed(seed)
    Faker.seed(seed)

    companies = []

    # 行业分布：零售业40%，批发业30%，餐饮业20%，住宿业10%
    industry_distribution = {
        'retail': int(count * 0.40),
        'wholesale': int(count * 0.30),
        'catering': int(count * 0.20),
        'accommodation': count - int(count * 0.40) - int(count * 0.30) - int(count * 0.20),
    }

    company_id = 1
    for industry, num in industry_distribution.items():
        for _ in range(num):
            # 生成企业名称
            prefix = random.choice(COMPANY_PREFIXES[industry])
            city = fake.city()
            suffix = random.choice(COMPANY_SUFFIXES)
            name = f"{city}{prefix}{suffix}"

            # 行业代码
            industry_code = random.choice(INDUSTRY_CODES[industry])

            # 企业规模 (1-大型, 2-中型, 3-小型, 4-微型)
            # 分布：大型5%，中型15%，小型40%，微型40%
            scale = random.choices([1, 2, 3, 4], weights=[5, 15, 40, 40])[0]

            # 根据规模生成营业额范围（万元）
            if scale == 1:  # 大型企业
                base_monthly = random.uniform(5000, 20000)
            elif scale == 2:  # 中型企业
                base_monthly = random.uniform(1000, 5000)
            elif scale == 3:  # 小型企业
                base_monthly = random.uniform(200, 1000)
            else:  # 微型企业
                base_monthly = random.uniform(50, 200)

            # 生成上年同期数据
            last_year_month = round(base_monthly * random.uniform(0.8, 1.2), 2)
            # 本期数据（有一定增长或下降）
            growth_rate = random.uniform(-0.15, 0.25)  # -15% 到 +25% 的增速
            current_month = round(last_year_month * (1 + growth_rate), 2)

            # 累计数据（假设当前是第6个月）
            cumulative_multiplier = random.uniform(5.5, 6.5)  # 约6个月的累计
            last_year_cumulative = round(last_year_month * cumulative_multiplier, 2)
            current_cumulative = round(current_month * cumulative_multiplier * random.uniform(0.95, 1.05), 2)

            # 销售额 >= 零售额
            sales_multiplier = random.uniform(1.0, 1.3)  # 销售额可能比零售额高
            sales_current_month = round(current_month * sales_multiplier, 2)
            sales_last_year_month = round(last_year_month * sales_multiplier, 2)
            sales_current_cumulative = round(current_cumulative * sales_multiplier, 2)
            sales_last_year_cumulative = round(last_year_cumulative * sales_multiplier, 2)

            companies.append({
                'id': company_id,
                'name': name,
                'credit_code': generate_credit_code(),
                'industry_code': industry_code,
                'scale': scale,
                'retail_current_month': current_month,
                'retail_last_year_month': last_year_month,
                'retail_current_cumulative': current_cumulative,
                'retail_last_year_cumulative': last_year_cumulative,
                'sales_current_month': sales_current_month,
                'sales_last_year_month': sales_last_year_month,
                'sales_current_cumulative': sales_current_cumulative,
                'sales_last_year_cumulative': sales_last_year_cumulative,
            })

            company_id += 1

    # 打乱顺序
    random.shuffle(companies)
    return companies


def create_excel_file(companies: list, output_path: str) -> str:
    """
    创建 Excel 文件

    Args:
        companies: 企业数据列表
        output_path: 输出文件路径

    Returns:
        创建的文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "企业数据"

    # 表头（与 FieldMapping 对应）
    headers = [
        '企业名称',
        '统一社会信用代码',
        '行业代码',
        '企业规模',
        '本期零售额',
        '上年同期零售额',
        '本年累计零售额',
        '上年累计零售额',
        '本期销售额',
        '上年同期销售额',
        '本年累计销售额',
        '上年累计销售额',
    ]

    # 写入表头
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 写入数据
    for row_idx, company in enumerate(companies, 2):
        ws.cell(row=row_idx, column=1, value=company['name'])
        ws.cell(row=row_idx, column=2, value=company['credit_code'])
        ws.cell(row=row_idx, column=3, value=company['industry_code'])
        ws.cell(row=row_idx, column=4, value=company['scale'])
        ws.cell(row=row_idx, column=5, value=company['retail_current_month'])
        ws.cell(row=row_idx, column=6, value=company['retail_last_year_month'])
        ws.cell(row=row_idx, column=7, value=company['retail_current_cumulative'])
        ws.cell(row=row_idx, column=8, value=company['retail_last_year_cumulative'])
        ws.cell(row=row_idx, column=9, value=company['sales_current_month'])
        ws.cell(row=row_idx, column=10, value=company['sales_last_year_month'])
        ws.cell(row=row_idx, column=11, value=company['sales_current_cumulative'])
        ws.cell(row=row_idx, column=12, value=company['sales_last_year_cumulative'])

    # 调整列宽
    column_widths = [30, 25, 12, 10, 15, 15, 15, 15, 15, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # 保存文件
    wb.save(output_path)
    return output_path


def get_field_mapping() -> dict:
    """返回 Excel 列名到 API 字段的映射"""
    return {
        "companyName": "企业名称",
        "creditCode": "统一社会信用代码",
        "industryCode": "行业代码",
        "companyScale": "企业规模",
        "retailCurrentMonth": "本期零售额",
        "retailLastYearMonth": "上年同期零售额",
        "retailCurrentCumulative": "本年累计零售额",
        "retailLastYearCumulative": "上年累计零售额",
        "salesCurrentMonth": "本期销售额",
        "salesLastYearMonth": "上年同期销售额",
        "salesCurrentCumulative": "本年累计销售额",
        "salesLastYearCumulative": "上年累计销售额",
    }


if __name__ == '__main__':
    # 生成测试数据
    companies = generate_company_data(count=300, seed=42)

    # 创建输出目录
    output_dir = os.path.join(os.path.dirname(__file__), 'fixtures')
    os.makedirs(output_dir, exist_ok=True)

    # 生成文件
    output_path = os.path.join(output_dir, 'test_companies_300.xlsx')
    create_excel_file(companies, output_path)

    print(f"已生成测试数据文件: {output_path}")
    print(f"共 {len(companies)} 条企业数据")

    # 统计各行业数量
    from collections import Counter
    industry_count = Counter()
    for c in companies:
        code = c['industry_code'][:2]
        if code == '51':
            industry_count['批发业'] += 1
        elif code == '52':
            industry_count['零售业'] += 1
        elif code == '61':
            industry_count['住宿业'] += 1
        elif code == '62':
            industry_count['餐饮业'] += 1

    print("\n行业分布:")
    for industry, count in industry_count.items():
        print(f"  {industry}: {count}")
